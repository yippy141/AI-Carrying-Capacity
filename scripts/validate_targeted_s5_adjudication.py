#!/usr/bin/env python3
"""Validate issue #41's bounded S5 adjudication without granting approval."""
from __future__ import annotations

import csv
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

import validate_fusion_domain_review as domain

ROOT = Path(__file__).resolve().parents[1]
PACKAGE = "research/structural-profiles-pilot/adjudication"
RECONCILIATION = "research/structural-profiles-pilot/reconciliation"
DOMAIN_PACKAGE = "research/structural-profiles-pilot/domain-review"
ADJUDICATION_FILE = "targeted_s5_adjudication_v1.csv"
PLAN_FILE = "three_anchor_human_review_plan_v1.csv"
NOTE_FILE = "S5_ADJUDICATION_NOTE.md"
WORKBOOK_FILE = "targeted_s5_adjudication_v1.xlsx"
DOMAIN_REVIEW_DIGEST = "243c2512caa15e709c1453b257bf3c7b05e3460cf4ce57c749a1566c43dd116c"

ADJUDICATION_FIELDS = """profile_id stage_id workflow sector pathway_id
seed_review_id seed_s5 seed_rationale seed_source_ids independent_review_id
independent_s5 independent_rationale independent_source_ids comparison_status
owner_exception_id owner_disposition owner_rationale owner_decision_state
domain_review_recommendation domain_review_confidence domain_review_disposition
selected_s5 recommended_low recommended_high adjudication_outcome
adjudication_confidence next_independent_boundary boundary_independence_status
boundary_independence_notes bounded_direct_consequence excluded_later_consequences
source_ids source_gap unresolved_gap_ids human_expert_package_id
canonical_approval_blocker draft_use_blocker draft_use_status coding_as_of
last_reviewed revisit_triggers adjudicated_by adjudicator_model reasoning_effort
notes""".split()
PLAN_FIELDS = """expert_package_id sector theme profile_ids dimension_cells
review_questions why_load_bearing source_ids requested_expertise named_reviewer
status open_gap_ids open_gap_status canonical_approval_blocker draft_use_blocker
draft_use_status pursuit_timing notes""".split()
VIEW_FIELDS = """profile_id workflow sector seed_s5 independent_s5
owner_disposition domain_review_recommendation selected_s5 recommended_low
recommended_high adjudication_outcome adjudication_confidence
next_independent_boundary bounded_direct_consequence excluded_later_consequences
source_ids source_gap unresolved_gap_ids human_expert_package_id
canonical_approval_blocker draft_use_blocker draft_use_status notes""".split()
SHEETS = ["Start Here", "Adjudication", "Software", "Manufacturing", "Fusion", "Unresolved", "Expert Packages"]
OUTCOMES = {
    "selected_provisional", "selected_provisional_pending_named_expert",
    "preserved_range", "preserved_disagreement", "needs_better_evidence",
}
EXPECTED = {
    "sp-0001": ("3", "", "", "selected_provisional_pending_named_expert", "medium", "EXP-SW-01"),
    "sp-0002": ("3", "", "", "selected_provisional_pending_named_expert", "medium", "EXP-SW-01"),
    "sp-0003": ("2", "", "", "selected_provisional_pending_named_expert", "medium", "EXP-SW-01"),
    "sp-0004": ("1", "", "", "selected_provisional_pending_named_expert", "medium", "EXP-SW-01"),
    "sp-0005": ("1", "", "", "selected_provisional_pending_named_expert", "medium", "EXP-SW-01"),
    "sp-0008": ("2", "", "", "selected_provisional_pending_named_expert", "medium", "EXP-MFG-01"),
    "sp-0011": ("", "1", "2", "preserved_range", "low", "EXP-MFG-01"),
    "sp-0012": ("", "1", "2", "preserved_range", "low", "EXP-MFG-01"),
    "sp-0013": ("", "1", "2", "preserved_range", "low", "EXP-MFG-01"),
    "sp-0014": ("", "2", "3", "preserved_range", "low", ""),
    "sp-0015": ("3", "", "", "selected_provisional", "medium", ""),
    "sp-0020": ("", "0", "1", "preserved_range", "low", "EXP-FUS-02"),
    "sp-0023": ("", "1", "2", "preserved_range", "low", "EXP-FUS-02"),
    "sp-0024": ("1", "", "", "selected_provisional_pending_named_expert", "low", "EXP-FUS-03"),
    "sp-0025": ("1", "", "", "selected_provisional_pending_named_expert", "low", "EXP-FUS-03"),
    "sp-0028": ("1", "", "", "selected_provisional_pending_named_expert", "low", "EXP-FUS-04"),
    "sp-0029": ("1", "", "", "selected_provisional_pending_named_expert", "low", "EXP-FUS-04"),
    "sp-0030": ("", "0", "2", "preserved_disagreement", "low", "EXP-FUS-05"),
    "sp-0031": ("1", "", "", "selected_provisional_pending_named_expert", "low", "EXP-FUS-06"),
}
FUSION_PACKAGES = {
    "EXP-FUS-01": ("gap-01",),
    "EXP-FUS-02": ("gap-02", "gap-03"),
    "EXP-FUS-03": ("gap-04", "gap-05"),
    "EXP-FUS-04": ("gap-06", "gap-07"),
    "EXP-FUS-05": ("gap-08",),
    "EXP-FUS-06": ("gap-09",),
}
UNRESOLVED = {"sp-0011", "sp-0012", "sp-0013", "sp-0014", "sp-0020", "sp-0023", "sp-0030"}
PROVENANCE = {
    "boundary_independence_status": "assumed_not_verified",
    "draft_use_blocker": "none",
    "draft_use_status": "allowed_for_private_use_staged_wp2_and_public_pilot_only_as_EXPERT-CODED_DRAFT_with_range_confidence_gaps_visible",
    "coding_as_of": "2026-09-03", "last_reviewed": "2026-09-03",
    "adjudicated_by": "Codex Desktop", "adjudicator_model": "gpt-5.6-sol",
    "reasoning_effort": "xhigh",
}


class S5AdjudicationValidationError(ValueError):
    """The issue #41 result or its protected-state contract failed."""


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        header = reader.fieldnames or []
        if not header or len(header) != len(set(header)):
            raise S5AdjudicationValidationError(f"{path.name}: empty or duplicate CSV header")
        rows = list(reader)
        if any(None in row or any(value is None for value in row.values()) for row in rows):
            raise S5AdjudicationValidationError(f"{path.name}: malformed CSV row")
        return rows


def _unique(values: list[str]) -> str:
    return ";".join(dict.fromkeys(value for value in values if value))


def validate_protected_inputs(root: Path = ROOT) -> list[str]:
    errors = domain.validate_protected_inputs(root)
    if domain.tree_digest(root, DOMAIN_PACKAGE) != DOMAIN_REVIEW_DIGEST:
        errors.append("Protected merged fusion domain-review package changed")
    return errors


def validate_records(rows: list[dict[str, str]], root: Path = ROOT) -> list[str]:
    errors: list[str] = []

    def check(condition: bool, message: str) -> None:
        if not condition:
            errors.append(message)

    check(all(list(row) == ADJUDICATION_FIELDS for row in rows),
          "Adjudication exact schema required; no composite, score, or approval columns")
    check(len(rows) == 19, "Exactly 19 targeted S5 adjudication rows required")
    check(len({row.get('profile_id') for row in rows}) == len(rows), "Duplicate S5 profile row")
    if errors:
        return errors

    backlog = read_rows(root / RECONCILIATION / "targeted_s5_adjudication_backlog_v1.csv")
    seed = {row["profile_id"]: row for row in read_rows(root / RECONCILIATION / "seed_submission_v1.csv")}
    independent = {row["profile_id"]: row for row in read_rows(root / RECONCILIATION / "independent_submission_v1.csv")}
    owners = {row["profile_id"]: row for row in read_rows(root / RECONCILIATION / "owner_decisions_v1.csv") if row["s_dimension"] == "S5"}
    fusion = {row["profile_id"]: row for row in read_rows(root / DOMAIN_PACKAGE / "fusion_dimension_review_v1.csv") if row["dimension"] == "S5"}
    sources = {row["source_id"]: row for row in read_rows(root / "data/sources/source_register.csv")}

    check([row["profile_id"] for row in rows] == [row["profile_id"] for row in backlog],
          "Wrong, reordered, missing, or added S5 backlog population")
    check(set(EXPECTED) == {row["profile_id"] for row in backlog}, "Expected 19-row decision map drifted from frozen backlog")
    if {row["profile_id"] for row in rows} != set(EXPECTED):
        return errors

    for row, base in zip(rows, backlog):
        profile = row["profile_id"]
        for field in ("profile_id", "stage_id", "workflow", "sector", "pathway_id", "comparison_status"):
            check(row[field] == base[field], f"{profile}: frozen backlog field changed ({field})")
        for prefix, source in (("seed", seed[profile]), ("independent", independent[profile])):
            expected_fields = {
                f"{prefix}_review_id": source["review_id"],
                f"{prefix}_s5": source["S5"],
                f"{prefix}_rationale": source["rationale"],
                f"{prefix}_source_ids": source["source_ids"],
            }
            for field, expected in expected_fields.items():
                check(row[field] == expected, f"{profile}: original {prefix} value/rationale/source not preserved ({field})")
        owner = owners.get(profile, {})
        expected_owner = {
            "owner_exception_id": owner.get("exception_id", ""),
            "owner_disposition": owner.get("owner_disposition", ""),
            "owner_rationale": owner.get("owner_rationale", ""),
            "owner_decision_state": owner.get("decision_state", ""),
        }
        for field, expected in expected_owner.items():
            check(row[field] == expected, f"{profile}: owner decision or rationale not preserved ({field})")

        domain_row = fusion.get(profile)
        if domain_row:
            recommendation = domain_row["recommended_low"] if domain_row["recommended_low"] == domain_row["recommended_high"] else f"{domain_row['recommended_low']}-{domain_row['recommended_high']}"
            for field, expected in {
                "domain_review_recommendation": recommendation,
                "domain_review_confidence": domain_row["recommendation_confidence"],
                "domain_review_disposition": domain_row["disposition"],
                "source_ids": "" if domain_row["source_ids"] == "missing" else domain_row["source_ids"],
                "unresolved_gap_ids": "" if domain_row["unresolved_gap_ids"] == "none" else domain_row["unresolved_gap_ids"],
            }.items():
                check(row[field] == expected, f"{profile}: merged fusion S5 record not preserved ({field})")
        else:
            check(not any(row[field] for field in ("domain_review_recommendation", "domain_review_confidence", "domain_review_disposition", "source_ids", "unresolved_gap_ids")),
                  f"{profile}: non-fusion row gained unsupported domain or source data")

        actual = (row["selected_s5"], row["recommended_low"], row["recommended_high"],
                  row["adjudication_outcome"], row["adjudication_confidence"], row["human_expert_package_id"])
        check(actual == EXPECTED[profile], f"{profile}: bounded S5 result, confidence, or expert route changed")
        check(row["adjudication_outcome"] in OUTCOMES, f"{profile}: invalid adjudication outcome")
        for field in ("selected_s5", "recommended_low", "recommended_high"):
            check(row[field] == "" or row[field] in {"0", "1", "2", "3", "4"}, f"{profile}: {field} must be an integer ordinal 0-4 or missing")
        is_point = row["adjudication_outcome"].startswith("selected_")
        check((bool(row["selected_s5"]) and not row["recommended_low"] and not row["recommended_high"]) if is_point else (not row["selected_s5"] and bool(row["recommended_low"]) and bool(row["recommended_high"])),
              f"{profile}: point and range/disagreement forms must be mutually exclusive")
        if (not is_point and row["recommended_low"] in {"0", "1", "2", "3", "4"}
                and row["recommended_high"] in {"0", "1", "2", "3", "4"}):
            check(int(row["recommended_low"]) < int(row["recommended_high"]), f"{profile}: range must be ordered and non-degenerate")
        for field, expected in PROVENANCE.items():
            check(row[field] == expected, f"{profile}: false or missing qualitative/draft provenance ({field})")
        for field in ("next_independent_boundary", "boundary_independence_notes", "bounded_direct_consequence", "excluded_later_consequences", "source_gap", "revisit_triggers", "notes"):
            check(len(row[field].strip()) >= 35, f"{profile}: missing substantive bounded-S5 explanation ({field})")
        check("profile_population_and_owner_canonical_approval_not_performed" in row["canonical_approval_blocker"],
              f"{profile}: canonical approval blocker lost")
        check("approved" not in row["adjudication_outcome"] and "canonical" not in row["adjudication_outcome"],
              f"{profile}: adjudication falsely grants approval")
        for source_id in filter(None, row["source_ids"].split(";")):
            check(source_id in sources, f"{profile}: unknown canonical source ID {source_id}")
            if source_id in sources:
                check(sources[source_id]["review_status"] == "reviewed", f"{profile}: source ID {source_id} is not reviewed")

    check(sum(bool(row["owner_disposition"]) for row in rows) == 7, "Exactly seven owner-routed S5 decisions required")
    check(Counter(row["adjudication_outcome"] for row in rows) == Counter({
        "selected_provisional": 1, "selected_provisional_pending_named_expert": 11,
        "preserved_range": 6, "preserved_disagreement": 1,
    }), "Adjudication outcome counts changed")
    check(Counter(row["adjudication_confidence"] for row in rows) == Counter({"medium": 7, "low": 12}),
          "Adjudication confidence counts changed")
    check(Counter(row["sector"] for row in rows) == Counter({
        "Software engineering": 5, "Discrete manufacturing": 4, "Fusion, magnetic confinement": 10,
    }), "Three-anchor sector counts changed")
    check({row["profile_id"] for row in rows if not row["selected_s5"]} == UNRESOLVED,
          "Seven unresolved point-selection rows changed")
    check(sum(not row["source_ids"] for row in rows) == 11, "Exactly 11 S5 rows must retain blank source IDs")
    return errors


def validate_plan(rows: list[dict[str, str]], root: Path = ROOT) -> list[str]:
    errors: list[str] = []

    def check(condition: bool, message: str) -> None:
        if not condition:
            errors.append(message)

    check(all(list(row) == PLAN_FIELDS for row in rows), "Human-review plan exact schema required")
    check(len(rows) == 8 and [row.get("expert_package_id") for row in rows] == [*FUSION_PACKAGES, "EXP-SW-01", "EXP-MFG-01"],
          "Exactly six retained fusion plus two non-fusion expert packages required")
    if errors:
        return errors
    queue = read_rows(root / DOMAIN_PACKAGE / "fusion_human_expert_queue_v1.csv")
    plan = {row["expert_package_id"]: row for row in rows}
    for package_id, gaps in FUSION_PACKAGES.items():
        source_rows = [row for row in queue if row["expert_package_id"] == package_id]
        row = plan[package_id]
        expected = {
            "profile_ids": _unique([item["profile_id"] for item in source_rows]),
            "dimension_cells": ";".join(f"{item['profile_id']}/{item['dimension']}" for item in source_rows),
            "review_questions": " | ".join(item["question_for_expert"] for item in source_rows),
            "why_load_bearing": " | ".join(item["why_load_bearing"] for item in source_rows),
            "source_ids": _unique([source for item in source_rows for source in ([] if item["source_ids"] == "missing" else item["source_ids"].split(";"))]) or "missing",
            "requested_expertise": _unique([item["requested_expertise"] for item in source_rows]),
            "open_gap_ids": ";".join(gaps),
        }
        for field, value in expected.items():
            check(row[field] == value, f"{package_id}: retained fusion questions/evidence/gaps changed ({field})")
        check(row["notes"].startswith("Retained unchanged"), f"{package_id}: retained-package status lost")

    check(sum(len([part for part in row["dimension_cells"].split(";") if part]) for row in rows[:6]) == 19,
          "All 19 exact fusion cell questions must remain in six packages")
    check(set(filter(None, ";".join(row["open_gap_ids"] for row in rows[:6]).split(";"))) == {f"gap-{number:02d}" for number in range(1, 10)},
          "Nine fusion empirical gaps must remain open")
    for package_id, profiles, cells, sector in (
        ("EXP-SW-01", "sp-0001;sp-0002;sp-0003;sp-0004;sp-0005", "sp-0001/S5;sp-0002/S5;sp-0003/S5;sp-0004/S5;sp-0005/S5", "Software engineering"),
        ("EXP-MFG-01", "sp-0008;sp-0011;sp-0012;sp-0013", "sp-0008/S5;sp-0011/S5;sp-0012/S5;sp-0013/S5", "Discrete manufacturing"),
    ):
        row = plan[package_id]
        check((row["profile_ids"], row["dimension_cells"], row["sector"], row["source_ids"]) == (profiles, cells, sector, "missing"),
              f"{package_id}: new non-fusion package scope or missing source state changed")
        check("next genuinely independent" in row["review_questions"] and len(row["why_load_bearing"]) > 100,
              f"{package_id}: boundary-independence review question is not substantive")
    for row in rows:
        package_id = row["expert_package_id"]
        check(row["named_reviewer"] == "missing" and row["status"] == "pending_named_specialist",
              f"{package_id}: package falsely records completed human review")
        check(row["open_gap_status"] in {"open", "not_applicable"}, f"{package_id}: empirical gap falsely closed")
        check(row["canonical_approval_blocker"] == "true" and row["draft_use_blocker"] == "false",
              f"{package_id}: canonical and draft-use blockers must remain distinct")
        check("EXPERT-CODED · DRAFT" in row["draft_use_status"] and "before canonical approval" in row["pursuit_timing"] and "not before staged WP2" in row["pursuit_timing"],
              f"{package_id}: labelled-draft or pursuit-timing contract changed")
    return errors


def validate_note(note: str, records: list[dict[str, str]], plan: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    required = [
        "19/19", "| `selected_provisional` | 1 |",
        "| `selected_provisional_pending_named_expert` | 11 |",
        "| `preserved_range` | 6 |", "| `preserved_disagreement` | 1 |",
        "| `needs_better_evidence` | 0 |", "| `medium` | 7 |", "| `low` | 12 |",
        "Eleven adjudication rows have blank `source_ids`", "gap-01", "gap-09",
        "All eight packages are pursued **before canonical approval**, not before",
        "No midpoint was manufactured", "No profile is approved or canonical, and WP2 has not begun",
    ]
    for text in required:
        if text.lower() not in note.lower():
            errors.append(f"S5 note missing required count, gap, or non-approval statement: {text}")
    for profile in UNRESOLVED:
        if profile not in note:
            errors.append(f"S5 note missing unresolved row {profile}")
    for package in (row["expert_package_id"] for row in plan):
        if package not in note:
            errors.append(f"S5 note missing expert package {package}")
    return errors


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sheet_rows(sheet, headers: list[str], count: int) -> list[dict[str, str]]:
    actual_headers = [_cell_text(sheet.cell(4, column).value) for column in range(1, len(headers) + 1)]
    if actual_headers != headers:
        return []
    return [{header: _cell_text(sheet.cell(row, column).value) for column, header in enumerate(headers, 1)}
            for row in range(5, count + 5)]


def validate_workbook(path: Path, records: list[dict[str, str]], plan: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    try:
        workbook = load_workbook(path, data_only=False)
    except Exception as exc:  # pragma: no cover - malformed ZIP path
        return [f"Workbook cannot be opened: {exc}"]
    if workbook.sheetnames != SHEETS:
        errors.append("Workbook must contain exactly the seven required visible sheets in order")
        return errors
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            errors.append(f"Workbook has hidden sheet: {sheet.title}")
        if any(dimension.hidden for dimension in sheet.row_dimensions.values()) or any(dimension.hidden for dimension in sheet.column_dimensions.values()):
            errors.append(f"Workbook has hidden rows or columns: {sheet.title}")
    allowed_formulas = {
        "B5": "=COUNTA('Adjudication'!$A$5:$A$23)",
        "B6": '=COUNTIF(\'Adjudication\'!$V$5:$V$23,"<>")',
        "B7": "=B5-B6",
        "B8": '=COUNTIF(\'Adjudication\'!$D$5:$D$23,"Software engineering")',
        "B9": '=COUNTIF(\'Adjudication\'!$D$5:$D$23,"Discrete manufacturing")',
        "B10": '=COUNTIF(\'Adjudication\'!$D$5:$D$23,"Fusion, magnetic confinement")',
        "B11": "=COUNTA('Expert Packages'!$A$5:$A$12)",
    }
    formulas: dict[str, str] = {}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    key = cell.coordinate if sheet.title == "Start Here" else f"{sheet.title}!{cell.coordinate}"
                    formulas[key] = f"={cell.value}" if not str(cell.value).startswith("=") else str(cell.value)
    if formulas != allowed_formulas:
        errors.append("Workbook contains a missing or unapproved formula/composite")
    detail = _sheet_rows(workbook["Adjudication"], ADJUDICATION_FIELDS, 19)
    if detail != records:
        errors.append("Workbook/CSV content mismatch on complete adjudication record")
    expected_views = {
        "Software": [row for row in records if row["sector"] == "Software engineering"],
        "Manufacturing": [row for row in records if row["sector"] == "Discrete manufacturing"],
        "Fusion": [row for row in records if row["sector"] == "Fusion, magnetic confinement"],
        "Unresolved": [row for row in records if not row["selected_s5"]],
    }
    for name, rows in expected_views.items():
        actual = _sheet_rows(workbook[name], VIEW_FIELDS, len(rows))
        projected = [{field: row[field] for field in VIEW_FIELDS} for row in rows]
        if actual != projected:
            errors.append(f"Workbook/CSV content mismatch on {name} view")
    actual_plan = _sheet_rows(workbook["Expert Packages"], PLAN_FIELDS, 8)
    if actual_plan != plan:
        errors.append("Workbook/CSV content mismatch on expert package plan")
    for sheet in workbook.worksheets[1:]:
        header_row = 4
        expected_rows = {"Adjudication": 19, "Software": 5, "Manufacturing": 4, "Fusion": 10, "Unresolved": 7, "Expert Packages": 8}[sheet.title]
        if any(_cell_text(sheet.cell(row, 1).value) for row in range(header_row + expected_rows + 1, sheet.max_row + 1)):
            errors.append(f"Workbook has extra static rows or composite drift: {sheet.title}")
    return errors


def validate_control_plane(root: Path = ROOT) -> list[str]:
    errors: list[str] = []
    files = {
        "project": root / "docs/PROJECT_STATE.md", "roadmap": root / "docs/ROADMAP.md",
        "status": root / "reports/PM_STATUS.md", "tasks": root / "docs/TASKS.md",
        "decisions": root / "docs/DECISIONS.md", "brief": root / "docs/AGENT_BRIEF.md",
        "readme": root / "README.md",
    }
    if missing := [name for name, path in files.items() if not path.is_file()]:
        return [f"Missing control-plane file(s): {', '.join(missing)}"]
    text = {name: path.read_text(encoding="utf-8") for name, path in files.items()}
    required_project = [
        "Frontier Is Not Fate", "AI Conversion Atlas", "frontier access", "conversion capacity",
        "adaptation capacity", "distribution quality", "realized outcomes", "accessible capability",
        "adjustment costs", "docs/AUTHORITATIVE_DOCS.md", "supersed", "31", "155", "23", "28",
        "44", "90", "19", "9 open", "software", "manufacturing", "fusion",
        "one substantive PM review", "one bounded correction pass", "P0/P1", "WP2",
    ]
    for phrase in required_project:
        if phrase.lower() not in text["project"].lower():
            errors.append(f"PROJECT_STATE missing current method/count/sequence phrase: {phrase}")
    for name in ("brief", "readme", "roadmap", "status"):
        if "docs/PROJECT_STATE.md" not in text[name]:
            errors.append(f"{name} missing PROJECT_STATE entry pointer")
    for phase in range(9):
        if f"Phase {phase}" not in text["roadmap"]:
            errors.append(f"ROADMAP missing Phase {phase}")
    for phrase in ("12 provisional points", "six preserved ranges", "one preserved disagreement", "retaining six fusion packages", "EXP-SW-01", "EXP-MFG-01"):
        if phrase.lower() not in text["decisions"].lower() and phrase.lower() not in text["tasks"].lower():
            errors.append(f"TASKS/DECISIONS missing issue #41 completion: {phrase}")
    if not re.search(r"19/19|19-row", text["status"]):
        errors.append("PM_STATUS missing exact targeted S5 coverage")
    if "WP2 has not begun" not in text["status"]:
        errors.append("PM_STATUS must state that WP2 has not begun")
    return errors


def validate(root: Path = ROOT) -> dict[str, object]:
    protected = validate_protected_inputs(root)
    package = root / PACKAGE
    records = read_rows(package / ADJUDICATION_FILE)
    plan = read_rows(package / PLAN_FILE)
    errors = protected
    errors.extend(validate_records(records, root))
    errors.extend(validate_plan(plan, root))
    errors.extend(validate_note((package / NOTE_FILE).read_text(encoding="utf-8"), records, plan))
    errors.extend(validate_workbook(package / WORKBOOK_FILE, records, plan))
    errors.extend(validate_control_plane(root))
    if errors:
        raise S5AdjudicationValidationError("\n- " + "\n- ".join(errors))
    return {
        "rows": len(records),
        "outcomes": dict(Counter(row["adjudication_outcome"] for row in records)),
        "confidence": dict(Counter(row["adjudication_confidence"] for row in records)),
        "point_form": sum(bool(row["selected_s5"]) for row in records),
        "range_or_disagreement_form": sum(not row["selected_s5"] for row in records),
        "expert_packages": len(plan), "open_fusion_gaps": 9, "canonical_approvals": 0,
    }


def main() -> None:
    try:
        result = validate()
    except (S5AdjudicationValidationError, domain.DomainReviewValidationError) as exc:
        print(f"Targeted S5 adjudication validation failed: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
    print("Targeted S5 adjudication validation passed")
    print(result)


if __name__ == "__main__":
    main()
