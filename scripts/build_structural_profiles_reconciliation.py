#!/usr/bin/env python3
"""Build the pinned Structural Profiles reconciliation package for issue #31."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import subprocess
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_structural_profiles_workbooks import RUBRIC_ROWS


ROOT = Path(__file__).resolve().parents[1]
WORKSHEET_DIR = ROOT / "research" / "structural-profiles-pilot" / "worksheet"
OUTPUT_DIR = ROOT / "research" / "structural-profiles-pilot" / "reconciliation"
AUDIT_ROOT = Path("/tmp/ai-capacity-recon-31-20260830")

BASE_SHA = "c2a5c53d586cecad4d137d459b78d432b9104870"
SEED_INITIAL_SHA = "5d7a14349dc6b0aea7cd6621a4e5f11bdcb06c16"
SEED_HEAD_SHA = "3bc7e1d3e16d507fe7374ae66a3af6eace519ca4"
INDEPENDENT_INITIAL_SHA = "3fd91a53bb4834dfd7974001104e63799e2a25a2"
INDEPENDENT_HEAD_SHA = "86dbff4aab694d413a33d3c4a8b0d28047d73c2f"

SEED_HEAD_CSV_PATH = (
    "research/structural-profiles-pilot/submissions/seed/seed_submission_v1.csv"
)
SEED_INITIAL_CSV_PATH = (
    "research/structural-profiles-pilot/submissions/fable/fable_submission_v1.csv"
)
INDEPENDENT_CSV_PATH = (
    "research/structural-profiles-pilot/submissions/independent/"
    "independent_submission_v1.csv"
)

SEED_WORKBOOK = (
    AUDIT_ROOT
    / "pr33/research/structural-profiles-pilot/submissions/seed/"
    "seed_submission_v1.xlsx"
)
INDEPENDENT_WORKBOOK = (
    AUDIT_ROOT
    / "pr32/research/structural-profiles-pilot/submissions/independent/"
    "independent_submission_v1.xlsx"
)

EXPECTED_RAW_HASHES = {
    "seed": "9723a3c3a006b701fed3d000f77d89379610bc51f85eff3555b696078708e675",
    "independent": "4176afa878a850d7155a95772884bec72756353440b588f30e5414a37acdb973",
}
EXPECTED_SCORE_HASHES = {
    "seed": "1c3666348097def3bb89328ca71007827d42cecfc0cba9073e3df9692d9fdd34",
    "independent": "915d4af9e7eb279b3926d97c541ecacb403936e822a7868ca1c2ae6820f2bc33",
}
EXPECTED_RATIONALE_HASHES = {
    "seed": "3b017d3b5add8914ab530a13c80cbdeabf4cb3910d8f0215b77289123cf25123",
    "independent": "5040d0a154571c847d0ae7a113dbed9eed3d68f669651416343b878d810b7d0c",
}

PROFILE_FIELDS = [
    "profile_id",
    "stage_id",
    "parent_stage_id",
    "sector",
    "workflow",
    "pathway_id",
    "application_context",
    "lifecycle_phase",
    "critical_path_role",
]
CANONICAL_REVIEW_FIELDS = [
    "review_id",
    "profile_id",
    "coder_type",
    "coder_role",
    "coder_name",
    "coder_model",
    "S1",
    "S2",
    "S3",
    "S4",
    "S5",
    "rationale",
    "source_ids",
    "coding_as_of",
    "submitted_at",
    "submission_status",
    "notes",
]
DIMENSIONS = ["S1", "S2", "S3", "S4", "S5"]
STATUSES = [
    "exact_agreement",
    "one_point_difference",
    "difference_ge_2",
    "missing_seed",
    "missing_independent",
    "missing_both",
]

SEED_PROVENANCE = {
    "coder_type": "model",
    "coder_role": "seed_proposer",
    "coder_name": "Claude Code",
    "coder_model": "claude-opus-5",
}
INDEPENDENT_PROVENANCE = {
    "coder_type": "model",
    "coder_role": "independent_coder",
    "coder_name": "Codex",
    "coder_model": "gpt-5.6",
}


def route(
    config: dict[tuple[str, str], dict[str, Any]],
    profile_id: str,
    dimensions: list[str],
    flags: list[str],
    reason: str,
) -> None:
    for dimension in dimensions:
        config[(profile_id, dimension)] = {
            "flags": list(flags),
            "reason": reason,
        }


OWNER_ROUTES: dict[tuple[str, str], dict[str, Any]] = {}
route(
    OWNER_ROUTES,
    "sp-0003",
    ["S5"],
    [
        "absolute_difference_ge_2",
        "contradictory_rationales",
        "missing_or_incompatible_source_support",
    ],
    "Two-point S5 dispute turns on whether escaped verification errors belong to "
    "this stage; the blank source base cannot resolve the convention.",
)
route(
    OWNER_ROUTES,
    "sp-0016",
    ["S2", "S3", "S4"],
    [
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
        "scope_pathway_application_or_lifecycle_ambiguity",
    ],
    "Experiment selection excludes physical execution but depends on scarce shots for "
    "feedback; one low-confidence coding makes this stage-boundary choice material.",
)
route(
    OWNER_ROUTES,
    "sp-0018",
    ["S1", "S2"],
    [
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
        "scope_pathway_application_or_lifecycle_ambiguity",
    ],
    "Plasma control combines digital development with live physical actuation; one "
    "low-confidence coding makes the digital/live weighting owner-relevant.",
)
route(
    OWNER_ROUTES,
    "sp-0020",
    ["S2", "S4"],
    ["missing_or_incompatible_source_support"],
    "Both coders assign the load-bearing zero values, but the conclusion depends on an "
    "uncited qualification-scale facility and exposure-time premise.",
)
route(
    OWNER_ROUTES,
    "sp-0024",
    ["S3", "S5"],
    [
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
    ],
    "Both coders report low confidence for scarce tritium-capable attempts and severe "
    "containment or inventory consequences, with no resolving source support.",
)
route(
    OWNER_ROUTES,
    "sp-0025",
    ["S2", "S5"],
    [
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
        "scope_pathway_application_or_lifecycle_ambiguity",
    ],
    "Both coders report low confidence and the blanket concept/pathway is not frozen "
    "tightly enough for evidence-free speed and error-tolerance disposition.",
)
route(
    OWNER_ROUTES,
    "sp-0028",
    ["S3", "S5"],
    [
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
    ],
    "Commissioning uses a unique integrated facility; a low-confidence coding and blank "
    "support materially affect attempt scarcity and consequence judgments.",
)
route(
    OWNER_ROUTES,
    "sp-0029",
    ["S5"],
    [
        "missing_or_incompatible_source_support",
        "scope_pathway_application_or_lifecycle_ambiguity",
    ],
    "The reliability target is not frozen, so the load-bearing S5 disagreement cannot be "
    "disposed without a scope choice or better evidence.",
)
route(
    OWNER_ROUTES,
    "sp-0030",
    ["S1", "S2", "S3", "S4"],
    [
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
        "scope_pathway_application_or_lifecycle_ambiguity",
    ],
    "The licensing row has low confidence and no jurisdiction; the scope changes the "
    "informational, timing, throughput, and physical-floor interpretation.",
)
route(
    OWNER_ROUTES,
    "sp-0030",
    ["S5"],
    [
        "absolute_difference_ge_2",
        "contradictory_rationales",
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
        "scope_pathway_application_or_lifecycle_ambiguity",
    ],
    "Two-point S5 dispute combines boundary allocation, low confidence, blank evidence, "
    "and an unspecified licensing jurisdiction.",
)
route(
    OWNER_ROUTES,
    "sp-0031",
    ["S3", "S4", "S5"],
    [
        "missing_or_incompatible_source_support",
        "low_confidence_load_bearing_stage",
        "scope_pathway_application_or_lifecycle_ambiguity",
    ],
    "Both coders report low confidence for an unprecedented, jurisdiction-sensitive "
    "pilot connection; blank evidence materially affects the three disputed dimensions.",
)


class ReconciliationBuildError(ValueError):
    """Raised when a pinned input or reconciliation invariant fails."""


def git_show(commit: str, path: str) -> bytes:
    try:
        return subprocess.check_output(["git", "show", f"{commit}:{path}"], cwd=ROOT)
    except subprocess.CalledProcessError as exc:
        raise ReconciliationBuildError(
            f"Unable to read pinned artifact {commit}:{path}."
        ) from exc


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def read_csv_bytes(data: bytes) -> list[dict[str, str]]:
    return list(csv.DictReader(io.StringIO(data.decode("utf-8"))))


def canonical_projection(
    rows: list[dict[str, str]], fields: list[str]
) -> list[list[str]]:
    return [[row[field] for field in fields] for row in rows]


def projection_hash(rows: list[dict[str, str]], fields: list[str]) -> str:
    payload = json.dumps(
        canonical_projection(rows, fields),
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return sha256(payload)


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        rendered = value.isoformat(timespec="milliseconds" if value.microsecond else "seconds")
        return rendered.removesuffix("+00:00") + "Z"
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    rendered = str(value)
    return rendered[1:] if rendered.startswith("'") else rendered


def read_profiles() -> list[dict[str, str]]:
    with (WORKSHEET_DIR / "stage_profiles_template.csv").open(
        newline="", encoding="utf-8"
    ) as handle:
        return list(csv.DictReader(handle))


def parse_sources(value: str) -> list[str]:
    if not value.strip():
        return []
    return sorted(
        {
            token.strip()
            for token in value.replace(";", "|").replace(",", "|").split("|")
            if token.strip()
        }
    )


def source_register_ids() -> set[str]:
    with (ROOT / "data" / "sources" / "source_register.csv").open(
        newline="", encoding="utf-8"
    ) as handle:
        return {row["source_id"] for row in csv.DictReader(handle)}


def validate_workbook_and_extract_confidence(
    workbook_path: Path,
    csv_rows: list[dict[str, str]],
    profiles: list[dict[str, str]],
    expected_provenance: dict[str, str],
) -> dict[str, str]:
    if not workbook_path.exists():
        raise ReconciliationBuildError(f"Pinned workbook is unavailable: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if workbook.sheetnames != ["S1_S5_RUBRIC", "SCOPE_REFERENCE", "SUBMISSION"]:
        raise ReconciliationBuildError(
            f"{workbook_path.name} has unexpected sheets: {workbook.sheetnames}"
        )
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            if any(cell.data_type == "f" for cell in row):
                raise ReconciliationBuildError(
                    f"{workbook_path.name} contains a prohibited formula."
                )

    submission = workbook["SUBMISSION"]
    headers = [normalize_cell(cell.value) for cell in submission[3]]
    workbook_rows = []
    for values in submission.iter_rows(min_row=4, values_only=True):
        row = {field: normalize_cell(value) for field, value in zip(headers, values)}
        if row.get("profile_id"):
            workbook_rows.append(row)
    if len(workbook_rows) != 31:
        raise ReconciliationBuildError(
            f"{workbook_path.name} must contain exactly 31 submission rows."
        )

    profile_by_id = {row["profile_id"]: row for row in profiles}
    csv_by_id = {row["profile_id"]: row for row in csv_rows}
    if [row["profile_id"] for row in workbook_rows] != list(csv_by_id):
        raise ReconciliationBuildError(
            f"{workbook_path.name} and CSV profile ordering differ."
        )

    confidence: dict[str, str] = {}
    for row in workbook_rows:
        profile_id = row["profile_id"]
        for field in PROFILE_FIELDS:
            if row[field] != profile_by_id[profile_id][field]:
                raise ReconciliationBuildError(
                    f"{workbook_path.name} scope drift at {profile_id} {field}."
                )
        for field in CANONICAL_REVIEW_FIELDS:
            if row[field] != csv_by_id[profile_id][field]:
                raise ReconciliationBuildError(
                    f"{workbook_path.name}/CSV mismatch at {profile_id} {field}."
                )
        if row["coding_confidence"] not in {"low", "medium", "high"}:
            raise ReconciliationBuildError(
                f"{workbook_path.name} missing confidence at {profile_id}."
            )
        confidence[profile_id] = row["coding_confidence"]
        for field, expected in expected_provenance.items():
            if row[field] != expected:
                raise ReconciliationBuildError(
                    f"{workbook_path.name} provenance mismatch at {profile_id} {field}."
                )
    return confidence


def validate_submission(
    label: str,
    rows: list[dict[str, str]],
    profiles: list[dict[str, str]],
    expected_provenance: dict[str, str],
) -> None:
    profile_ids = [row["profile_id"] for row in profiles]
    if len(rows) != 31 or [row["profile_id"] for row in rows] != profile_ids:
        raise ReconciliationBuildError(
            f"{label} submission must contain 31 profiles in canonical order."
        )
    if len(set(row["profile_id"] for row in rows)) != 31:
        raise ReconciliationBuildError(f"{label} submission contains duplicate profiles.")

    register = source_register_ids()
    forbidden = ("average", "composite", "percentage", "country_modifier", "C1")
    if any(any(token.lower() in field.lower() for token in forbidden) for field in rows[0]):
        raise ReconciliationBuildError(f"{label} submission contains derived fields.")

    for row in rows:
        profile_id = row["profile_id"]
        for field, expected in expected_provenance.items():
            if row[field] != expected:
                raise ReconciliationBuildError(
                    f"{label} provenance mismatch at {profile_id} {field}."
                )
        for dimension in DIMENSIONS:
            if row[dimension] not in {"0", "1", "2", "3", "4"}:
                raise ReconciliationBuildError(
                    f"{label} has invalid {dimension} at {profile_id}."
                )
        required = (
            "review_id",
            "rationale",
            "coding_as_of",
            "submitted_at",
            "submission_status",
            "notes",
        )
        if any(not row[field].strip() for field in required):
            raise ReconciliationBuildError(
                f"{label} has a missing required field at {profile_id}."
            )
        rationale_lower = row["rationale"].lower()
        for marker in ("s1", "s2", "s3 cost", "s3 throughput", "s4", "s5", "scope caveat"):
            if marker not in rationale_lower:
                raise ReconciliationBuildError(
                    f"{label} rationale omits {marker} at {profile_id}."
                )
        sources = parse_sources(row["source_ids"])
        unresolved = [source for source in sources if source not in register]
        if unresolved:
            raise ReconciliationBuildError(
                f"{label} has non-resolving source IDs at {profile_id}: {unresolved}."
            )
        if not sources and "source_gap:" not in row["notes"]:
            raise ReconciliationBuildError(
                f"{label} blank source_ids lacks a source_gap note at {profile_id}."
            )


def compare_initial_and_head(
    label: str,
    initial_rows: list[dict[str, str]],
    head_rows: list[dict[str, str]],
) -> None:
    score_fields = ["profile_id", *DIMENSIONS]
    rationale_fields = ["profile_id", "rationale"]
    score_hash = projection_hash(initial_rows, score_fields)
    rationale_hash = projection_hash(initial_rows, rationale_fields)
    if canonical_projection(initial_rows, score_fields) != canonical_projection(
        head_rows, score_fields
    ):
        raise ReconciliationBuildError(
            f"{label} scores changed after the initial submission; stop."
        )
    if canonical_projection(initial_rows, rationale_fields) != canonical_projection(
        head_rows, rationale_fields
    ):
        raise ReconciliationBuildError(
            f"{label} rationales changed after the initial submission; stop."
        )
    if score_hash != EXPECTED_SCORE_HASHES[label]:
        raise ReconciliationBuildError(f"{label} score projection hash drifted.")
    if rationale_hash != EXPECTED_RATIONALE_HASHES[label]:
        raise ReconciliationBuildError(f"{label} rationale projection hash drifted.")


def status_for(seed: int | None, independent: int | None) -> tuple[str, int | None]:
    if seed is None and independent is None:
        return "missing_both", None
    if seed is None:
        return "missing_seed", None
    if independent is None:
        return "missing_independent", None
    difference = abs(seed - independent)
    if difference == 0:
        return "exact_agreement", 0
    if difference == 1:
        return "one_point_difference", 1
    return "difference_ge_2", difference


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def submission_view(
    row: dict[str, str],
    dimension: str,
    confidence: str,
    reasoning_effort: str | None,
) -> dict[str, Any]:
    return {
        "review_id": row["review_id"],
        "coder_role": row["coder_role"],
        "coder_name": row["coder_name"],
        "coder_model": row["coder_model"],
        "reasoning_effort": reasoning_effort,
        "value": int(row[dimension]) if row[dimension] else None,
        "rationale": row["rationale"] or None,
        "source_ids": parse_sources(row["source_ids"]),
        "confidence": confidence,
    }


def build_note(
    generated_at: str,
    counts: Counter[str],
    exceptions: list[dict[str, Any]],
    profiles: list[dict[str, str]],
    seed_confidence: dict[str, str],
    independent_confidence: dict[str, str],
    s5_rows: list[dict[str, Any]],
) -> str:
    profile_by_id = {row["profile_id"]: row for row in profiles}
    trigger_counts = Counter(
        flag for row in exceptions for flag in row["semantic_flags"]
    )
    sector_counts = Counter(
        profile_by_id[row["profile_id"]]["sector"] for row in exceptions
    )
    seed_conf_counts = Counter(seed_confidence.values())
    independent_conf_counts = Counter(independent_confidence.values())

    lines = [
        "# Structural Profiles reconciliation — v1",
        "",
        "## Input provenance and stop-condition verification",
        "",
        f"- Branch base: `{BASE_SHA}` (`main` at branch creation).",
        f"- Seed submission: draft PR #33 head `{SEED_HEAD_SHA}`; CI completed successfully at 2026-08-30T14:25:05Z.",
        f"- Independent submission: draft PR #32 head `{INDEPENDENT_HEAD_SHA}`; CI completed successfully at 2026-08-30T14:29:20Z.",
        "- Seed coder: Claude Code / Claude Opus 5 (`claude-opus-5`), role `seed_proposer`.",
        "- Independent coder: Codex / GPT-5.6 (`gpt-5.6`), role `independent_coder`, reasoning effort `extra_high`.",
        "- Fable is credited as framework architect and is not credited as either row-level coder.",
        f"- Seed raw CSV SHA-256: `{EXPECTED_RAW_HASHES['seed']}`.",
        f"- Independent raw CSV SHA-256: `{EXPECTED_RAW_HASHES['independent']}`.",
        f"- Seed 155-value projection SHA-256: `{EXPECTED_SCORE_HASHES['seed']}`; rationale projection SHA-256: `{EXPECTED_RATIONALE_HASHES['seed']}`.",
        f"- Independent 155-value projection SHA-256: `{EXPECTED_SCORE_HASHES['independent']}`; rationale projection SHA-256: `{EXPECTED_RATIONALE_HASHES['independent']}`.",
        "- Field-level comparison against the first substantive commits confirms that both correction commits changed metadata only: no S-value or rationale changed.",
        "- Each workbook independently matched its CSV across all 17 canonical review fields; each had 31 profiles in frozen order, valid 0–4 integers, matching frozen scope, present provenance/rationales/dates/status/confidence, no formulas or prohibited derived fields, and blank-but-permitted source IDs.",
        "",
        "The two normalized CSVs are immutable byte copies read directly from those recorded heads. The seed-specific workflow normalization changes names and provenance labels only; it does not reinterpret or recode either submission.",
        "",
        "## Comparison counts",
        "",
        "| Comparison status | Count |",
        "| --- | ---: |",
    ]
    lines.extend(f"| `{status}` | {counts[status]} |" for status in STATUSES)
    lines.extend(
        [
            f"| **Total** | **{sum(counts.values())}** |",
            "",
            f"Owner-routed exceptions: **{len(exceptions)}**. Exact and one-point rows remain audit-only unless one of the recorded semantic triggers requires owner disposition. Trigger counts below are non-exclusive.",
            "",
            "| Owner-routing trigger | Exception rows |",
            "| --- | ---: |",
        ]
    )
    for trigger in sorted(trigger_counts):
        lines.append(f"| `{trigger}` | {trigger_counts[trigger]} |")
    lines.extend(
        [
            "",
            "| Sector | Owner exceptions |",
            "| --- | ---: |",
        ]
    )
    for sector in sorted(sector_counts):
        lines.append(f"| {sector} | {sector_counts[sector]} |")

    lines.extend(
        [
            "",
            "## Systematic source gap and confidence",
            "",
            "Both submissions have blank canonical `source_ids` on 31 of 31 profile rows (62 blank submission fields total, or 155 blank comparison views per coder). There are no non-resolving nonblank IDs. This systematic gap is counted here and is not itself replicated into 155 owner decisions. It routes only where the exception row records that the gap materially affects a load-bearing, low-confidence, scope-sensitive, or disputed judgment.",
            "",
            "| Coder | High | Medium | Low | Total profiles |",
            "| --- | ---: | ---: | ---: | ---: |",
            f"| Seed — Claude Code / Claude Opus 5 | {seed_conf_counts['high']} | {seed_conf_counts['medium']} | {seed_conf_counts['low']} | 31 |",
            f"| Independent — Codex / GPT-5.6, extra-high | {independent_conf_counts['high']} | {independent_conf_counts['medium']} | {independent_conf_counts['low']} | 31 |",
            "",
            "Confidence is preserved per coder and is not used to average, weight, or midpoint-select scores.",
            "",
            "## Cross-cutting coding convention — S5 boundary allocation",
            "",
            "The repeated S5 difference is a boundary-allocation question, not 19 separate instructions to select a preferred row. The seed coder generally assesses errors contained within the scoped stage and assigns escaped-error consequences downstream. The independent coder more often includes the reasonably foreseeable consequences of an erroneous stage output escaping into deployment, operations, qualification, or licensed operation.",
            "",
            "Owner question: should S5 assess **(a) locally contained errors only**, or **(b) the reasonably foreseeable consequences of an erroneous stage output escaping that stage**?",
            "",
            "No convention is selected here. The owner workbook leaves both the convention choice and post-reconciliation correction route blank. All original values remain unchanged.",
            "",
            f"S5 differs on {len(s5_rows)} profiles: seed is higher on 18, equal on 12 of all 31 profiles, and lower on 1. The two two-point gaps are `sp-0003` verification and validation and `sp-0030` licensing; there are no larger gaps.",
            "",
            "| Profile | Stage | Seed S5 | Independent S5 | Difference |",
            "| --- | --- | ---: | ---: | ---: |",
        ]
    )
    for row in s5_rows:
        profile = profile_by_id[row["profile_id"]]
        lines.append(
            f"| `{row['profile_id']}` | {profile['workflow']} | {row['seed']['value']} | {row['independent']['value']} | {row['numeric_difference']} |"
        )

    lines.extend(
        [
            "",
            "## Fusion domain-review queue",
            "",
            "`fusion_domain_review_queue_v1.csv` contains exactly 18 profiles (`sp-0014` through `sp-0031`). It is a routine domain-review queue, separate from owner exceptions. A fusion row appears in the owner subset only when another recorded semantic trigger applies.",
            "",
            "## Gate status and next step",
            "",
            "Structural blocker: none. The comparison denominator is 155 and the six status counts sum to 155. The owner-exception workbook contains no prefilled disposition, rationale, S5 convention choice, or correction route.",
            "",
            "Jinhua reviews `owner_exception_review_v1.xlsx` next. This package does not select, approve, canonicalize, or implement any profile row. It creates no country modifiers, governance codings, public UI, WP2 data, composite, average, midpoint, or forced consensus.",
            "",
            f"Generated at `{generated_at}`.",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    profiles = read_profiles()

    seed_bytes = git_show(SEED_HEAD_SHA, SEED_HEAD_CSV_PATH)
    independent_bytes = git_show(INDEPENDENT_HEAD_SHA, INDEPENDENT_CSV_PATH)
    if sha256(seed_bytes) != EXPECTED_RAW_HASHES["seed"]:
        raise ReconciliationBuildError("Pinned seed raw CSV hash drifted.")
    if sha256(independent_bytes) != EXPECTED_RAW_HASHES["independent"]:
        raise ReconciliationBuildError("Pinned independent raw CSV hash drifted.")

    seed_rows = read_csv_bytes(seed_bytes)
    independent_rows = read_csv_bytes(independent_bytes)
    seed_initial_rows = read_csv_bytes(git_show(SEED_INITIAL_SHA, SEED_INITIAL_CSV_PATH))
    independent_initial_rows = read_csv_bytes(
        git_show(INDEPENDENT_INITIAL_SHA, INDEPENDENT_CSV_PATH)
    )
    compare_initial_and_head("seed", seed_initial_rows, seed_rows)
    compare_initial_and_head(
        "independent", independent_initial_rows, independent_rows
    )
    validate_submission("seed", seed_rows, profiles, SEED_PROVENANCE)
    validate_submission(
        "independent", independent_rows, profiles, INDEPENDENT_PROVENANCE
    )

    seed_confidence = validate_workbook_and_extract_confidence(
        SEED_WORKBOOK, seed_rows, profiles, SEED_PROVENANCE
    )
    independent_confidence = validate_workbook_and_extract_confidence(
        INDEPENDENT_WORKBOOK,
        independent_rows,
        profiles,
        INDEPENDENT_PROVENANCE,
    )

    (OUTPUT_DIR / "seed_submission_v1.csv").write_bytes(seed_bytes)
    (OUTPUT_DIR / "independent_submission_v1.csv").write_bytes(independent_bytes)

    profile_by_id = {row["profile_id"]: row for row in profiles}
    seed_by_id = {row["profile_id"]: row for row in seed_rows}
    independent_by_id = {row["profile_id"]: row for row in independent_rows}
    rubric_by_dimension = {row["dimension"]: row for row in RUBRIC_ROWS}

    audit: list[dict[str, Any]] = []
    audit_csv: list[dict[str, Any]] = []
    exceptions: list[dict[str, Any]] = []
    owner_csv: list[dict[str, Any]] = []

    for profile in profiles:
        profile_id = profile["profile_id"]
        for dimension in DIMENSIONS:
            seed_view = submission_view(
                seed_by_id[profile_id],
                dimension,
                seed_confidence[profile_id],
                None,
            )
            independent_view = submission_view(
                independent_by_id[profile_id],
                dimension,
                independent_confidence[profile_id],
                "extra_high",
            )
            comparison_status, numeric_difference = status_for(
                seed_view["value"], independent_view["value"]
            )
            owner_config = OWNER_ROUTES.get((profile_id, dimension))
            semantic_flags = list(owner_config["flags"]) if owner_config else []
            if numeric_difference == 4 and "extreme_0_vs_4" not in semantic_flags:
                semantic_flags.append("extreme_0_vs_4")
            if numeric_difference is None and "missing_score" not in semantic_flags:
                semantic_flags.append("missing_score")
            owner_required = owner_config is not None or "missing_score" in semantic_flags
            fusion = profile["sector"] == "Fusion, magnetic confinement"
            if owner_required and fusion:
                review_route = "owner_and_domain_review"
            elif owner_required:
                review_route = "owner_review"
            elif fusion:
                review_route = "domain_review"
            else:
                review_route = "audit_only"

            row = {
                "profile_id": profile_id,
                "stage_id": profile["stage_id"],
                "pathway_id": profile["pathway_id"],
                "s_dimension": dimension,
                "comparison_status": comparison_status,
                "numeric_difference": numeric_difference,
                "seed": seed_view,
                "independent": independent_view,
                "semantic_flags": sorted(semantic_flags),
                "owner_review_required": owner_required,
                "review_route": review_route,
            }
            audit.append(row)

            audit_csv.append(
                {
                    "profile_id": profile_id,
                    "stage_id": profile["stage_id"],
                    "sector": profile["sector"],
                    "workflow": profile["workflow"],
                    "pathway_id": profile["pathway_id"],
                    "application_context": profile["application_context"],
                    "lifecycle_phase": profile["lifecycle_phase"],
                    "s_dimension": dimension,
                    "comparison_status": comparison_status,
                    "numeric_difference": "" if numeric_difference is None else numeric_difference,
                    "seed_review_id": seed_view["review_id"],
                    "seed_coder_name": seed_view["coder_name"],
                    "seed_coder_model": seed_view["coder_model"],
                    "seed_value": "" if seed_view["value"] is None else seed_view["value"],
                    "seed_rationale": seed_view["rationale"] or "",
                    "seed_source_ids": "|".join(seed_view["source_ids"]),
                    "seed_confidence": seed_view["confidence"] or "",
                    "independent_review_id": independent_view["review_id"],
                    "independent_coder_name": independent_view["coder_name"],
                    "independent_coder_model": independent_view["coder_model"],
                    "independent_reasoning_effort": independent_view["reasoning_effort"] or "",
                    "independent_value": "" if independent_view["value"] is None else independent_view["value"],
                    "independent_rationale": independent_view["rationale"] or "",
                    "independent_source_ids": "|".join(independent_view["source_ids"]),
                    "independent_confidence": independent_view["confidence"] or "",
                    "semantic_flags": "|".join(sorted(semantic_flags)),
                    "owner_review_required": str(owner_required).lower(),
                    "review_route": review_route,
                }
            )

            if owner_required:
                exception_id = f"exc-{profile_id}-{dimension.lower()}"
                exception = {
                    "exception_id": exception_id,
                    **row,
                    "owner_disposition": None,
                    "owner_rationale": None,
                    "status": "pending_owner",
                }
                exceptions.append(exception)
                rubric = rubric_by_dimension[dimension]
                owner_csv.append(
                    {
                        "exception_id": exception_id,
                        "profile_id": profile_id,
                        "stage_id": profile["stage_id"],
                        "workflow": profile["workflow"],
                        "sector": profile["sector"],
                        "pathway_id": profile["pathway_id"],
                        "application_context": profile["application_context"],
                        "lifecycle_phase": profile["lifecycle_phase"],
                        "s_dimension": dimension,
                        "dimension_name": rubric["name"],
                        "zero_endpoint": rubric["0_means"],
                        "two_guidance": rubric["2_guidance"],
                        "four_endpoint": rubric["4_means"],
                        "comparison_status": comparison_status,
                        "numeric_difference": "" if numeric_difference is None else numeric_difference,
                        "semantic_flags": "|".join(sorted(semantic_flags)),
                        "routing_reason": owner_config["reason"] if owner_config else "Missing score requires owner review.",
                        "seed_review_id": seed_view["review_id"],
                        "seed_coder_name": seed_view["coder_name"],
                        "seed_coder_model": seed_view["coder_model"],
                        "seed_value": "" if seed_view["value"] is None else seed_view["value"],
                        "seed_rationale": seed_view["rationale"] or "",
                        "seed_confidence": seed_view["confidence"] or "",
                        "seed_source_ids": "|".join(seed_view["source_ids"]),
                        "independent_review_id": independent_view["review_id"],
                        "independent_coder_name": independent_view["coder_name"],
                        "independent_coder_model": independent_view["coder_model"],
                        "independent_reasoning_effort": independent_view["reasoning_effort"] or "",
                        "independent_value": "" if independent_view["value"] is None else independent_view["value"],
                        "independent_rationale": independent_view["rationale"] or "",
                        "independent_confidence": independent_view["confidence"] or "",
                        "independent_source_ids": "|".join(independent_view["source_ids"]),
                        "review_route": review_route,
                        "owner_disposition": "",
                        "owner_rationale": "",
                        "status": "pending_owner",
                    }
                )

    counts = Counter(row["comparison_status"] for row in audit)
    if len(audit) != 155 or sum(counts.values()) != 155:
        raise ReconciliationBuildError("Comparison audit denominator is not 155.")
    if counts != Counter(
        {
            "exact_agreement": 96,
            "one_point_difference": 57,
            "difference_ge_2": 2,
        }
    ):
        raise ReconciliationBuildError(f"Unexpected comparison counts: {counts}")
    if len(exceptions) != 23:
        raise ReconciliationBuildError(
            f"Owner exception routing must produce 23 rows, found {len(exceptions)}."
        )

    fusion_queue_json = []
    fusion_queue_csv = []
    for profile in profiles:
        if profile["sector"] != "Fusion, magnetic confinement":
            continue
        profile_id = profile["profile_id"]
        seed_row = seed_by_id[profile_id]
        independent_row = independent_by_id[profile_id]
        fusion_queue_json.append(
            {
                "profile_id": profile_id,
                "stage_id": profile["stage_id"],
                "pathway_id": profile["pathway_id"],
                "reason": "fusion_requires_domain_informed_review",
                "owner_decision_required": False,
                "status": "pending",
                "domain_reviewer": None,
                "notes": None,
                "seed_review_id": seed_row["review_id"],
                "seed_coder_name": seed_row["coder_name"],
                "seed_coder_model": seed_row["coder_model"],
                "independent_review_id": independent_row["review_id"],
                "independent_coder_name": independent_row["coder_name"],
                "independent_coder_model": independent_row["coder_model"],
                "independent_reasoning_effort": "extra_high",
            }
        )
        fusion_queue_csv.append(
            {
                "profile_id": profile_id,
                "stage_id": profile["stage_id"],
                "parent_stage_id": profile["parent_stage_id"],
                "workflow": profile["workflow"],
                "pathway_id": profile["pathway_id"],
                "application_context": profile["application_context"],
                "lifecycle_phase": profile["lifecycle_phase"],
                "seed_review_id": seed_row["review_id"],
                "seed_coder_name": seed_row["coder_name"],
                "seed_coder_model": seed_row["coder_model"],
                **{f"seed_{dimension}": seed_row[dimension] for dimension in DIMENSIONS},
                "seed_confidence": seed_confidence[profile_id],
                "seed_source_ids": seed_row["source_ids"],
                "independent_review_id": independent_row["review_id"],
                "independent_coder_name": independent_row["coder_name"],
                "independent_coder_model": independent_row["coder_model"],
                "independent_reasoning_effort": "extra_high",
                **{
                    f"independent_{dimension}": independent_row[dimension]
                    for dimension in DIMENSIONS
                },
                "independent_confidence": independent_confidence[profile_id],
                "independent_source_ids": independent_row["source_ids"],
                "reason": "fusion_requires_domain_informed_review",
                "owner_decision_required": "false",
                "status": "pending",
                "domain_reviewer": "",
                "notes": "",
            }
        )
    if len(fusion_queue_json) != 18:
        raise ReconciliationBuildError("Fusion domain-review queue is not 18 profiles.")

    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace(
        "+00:00", "Z"
    )
    report = {
        "metadata": {
            "generated_at": generated_at,
            "seed_submission_id": f"pr-33@{SEED_HEAD_SHA}",
            "independent_submission_id": f"pr-32@{INDEPENDENT_HEAD_SHA}",
            "comparison_key": ["profile_id", "s_dimension"],
            "comparison_rule": "compare_seed_and_independent_by_profile_id_and_s_dimension",
            "expected_comparison_count": 155,
            "aggregation_rule": "preserve_submissions_no_averaging_midpoint_or_forced_consensus",
            "fusion_review_rule": "queue_every_fusion_profile_for_routine_domain_review",
        },
        "comparison_audit": audit,
        "summary_counts": {
            **{status: counts[status] for status in STATUSES},
            "total_comparisons": 155,
        },
        "exceptions": exceptions,
        "routine_domain_reviews": fusion_queue_json,
    }
    (OUTPUT_DIR / "reconciliation_report_v1.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    audit_headers = list(audit_csv[0])
    owner_headers = list(owner_csv[0])
    fusion_headers = list(fusion_queue_csv[0])
    write_csv(OUTPUT_DIR / "comparison_audit_v1.csv", audit_headers, audit_csv)
    write_csv(OUTPUT_DIR / "owner_exceptions_v1.csv", owner_headers, owner_csv)
    write_csv(
        OUTPUT_DIR / "fusion_domain_review_queue_v1.csv",
        fusion_headers,
        fusion_queue_csv,
    )

    s5_rows = [
        row
        for row in audit
        if row["s_dimension"] == "S5" and row["numeric_difference"]
    ]
    (OUTPUT_DIR / "RECONCILIATION_NOTE.md").write_text(
        build_note(
            generated_at,
            counts,
            exceptions,
            profiles,
            seed_confidence,
            independent_confidence,
            s5_rows,
        ),
        encoding="utf-8",
    )
    print(
        "Built reconciliation: "
        f"{len(audit)} audit rows, {len(exceptions)} owner exceptions, "
        f"{len(fusion_queue_json)} fusion reviews."
    )


if __name__ == "__main__":
    main()
