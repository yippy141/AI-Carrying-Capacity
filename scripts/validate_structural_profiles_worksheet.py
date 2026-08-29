#!/usr/bin/env python3
"""Validate the blank, owner-reviewable Structural Profiles worksheet package."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import tempfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

try:
    from scripts.build_structural_profiles_workbooks import (
        ALLOWED_VALUES,
        BLIND_WORKBOOK,
        COUNTRY_HEADERS,
        DEFAULT_PACKAGE_DIR,
        FABLE_WORKBOOK,
        GOVERNANCE_HEADERS,
        OWNER_HEADERS,
        PROFILE_HEADERS,
        REFERENCE_WORKBOOK,
        REVIEW_HEADERS,
        RUBRIC_HEADERS,
        RUBRIC_ROWS,
        SCOPE_REFERENCE_HEADERS,
        STAGE_HEADERS,
        SUBMISSION_HEADERS,
        build_all,
    )
except ModuleNotFoundError:  # Supports direct execution from scripts/.
    from build_structural_profiles_workbooks import (  # type: ignore[no-redef]
        ALLOWED_VALUES,
        BLIND_WORKBOOK,
        COUNTRY_HEADERS,
        DEFAULT_PACKAGE_DIR,
        FABLE_WORKBOOK,
        GOVERNANCE_HEADERS,
        OWNER_HEADERS,
        PROFILE_HEADERS,
        REFERENCE_WORKBOOK,
        REVIEW_HEADERS,
        RUBRIC_HEADERS,
        RUBRIC_ROWS,
        SCOPE_REFERENCE_HEADERS,
        STAGE_HEADERS,
        SUBMISSION_HEADERS,
        build_all,
    )


EXPECTED_LEAVES = [
    "requirements_and_architecture",
    "implementation",
    "verification_and_validation",
    "deployment",
    "operations_and_maintenance",
    "product_design",
    "process_engineering",
    "supply_and_tooling",
    "retooling_and_commissioning",
    "production_planning_and_scheduling",
    "production_and_process_control",
    "quality_assurance",
    "maintenance_and_continuous_improvement",
    "theory_and_system_design",
    "simulation",
    "experiment_selection",
    "diagnostics",
    "plasma_control",
    "materials_discovery_and_screening",
    "materials_qualification",
    "magnets",
    "heating_and_current_drive",
    "plasma_facing_components",
    "tritium_and_fuel_cycle",
    "blankets",
    "component_fabrication",
    "construction",
    "commissioning",
    "reliability_demonstration",
    "licensing",
    "grid_integration",
]

EXPECTED_FUSION_PARENTS = {
    "concept_and_design",
    "digital_experiment_loop",
    "plasma_operations",
    "materials",
    "plant_subsystems",
    "build",
    "demonstration_and_assurance",
    "approval_and_connection",
}

EXPECTED_PATHWAYS = {
    **{
        stage: "mature_software_delivery_and_maintenance"
        for stage in EXPECTED_LEAVES[:5]
    },
    **{
        stage: "discrete_manufacturing_npi_and_operations"
        for stage in EXPECTED_LEAVES[5:13]
    },
    **{
        stage: "tokamak_research_to_pilot_plant_demonstration"
        for stage in EXPECTED_LEAVES[13:]
    },
}

EXPECTED_PROFILE_IDS = [f"sp-{index:04d}" for index in range(1, 32)]
EXPECTED_LIFECYCLE_PHASES = dict(
    zip(
        EXPECTED_PROFILE_IDS,
        [
            "development",
            "development",
            "development",
            "commercial_deployment",
            "operations",
            "development",
            "development",
            "scale_up",
            "scale_up",
            "operations",
            "operations",
            "operations",
            "operations",
            "research",
            "research",
            "research",
            "research",
            "development",
            "research",
            "qualification",
            "development",
            "development",
            "development",
            "development",
            "development",
            "scale_up",
            "demonstration",
            "demonstration",
            "demonstration",
            "qualification",
            "demonstration",
        ],
        strict=True,
    )
)
S_FIELDS = ["S1", "S2", "S3", "S4", "S5"]
C_FIELDS = [field for field in COUNTRY_HEADERS if re.fullmatch(r"C[1-8]_.+", field)]
SCOPE_FIELDS = {
    "profile_id",
    "stage_id",
    "parent_stage_id",
    "sector",
    "workflow",
    "pathway_id",
    "application_context",
    "lifecycle_phase",
    "critical_path_role",
}

FABLE_IDENTITY = {
    "coder_type": "model",
    "coder_role": "seed_proposer",
    "coder_name": "fable",
    "coder_model": "claude-fable-5",
}

FORBIDDEN_BLIND_TEXT = ("fable", "claude-fable-5", "seed_proposer")
FORBIDDEN_DERIVED_HEADER_TOKENS = (
    "score",
    "average",
    "weighted",
    "composite",
    "percentage",
    "percent",
)

REFERENCE_SHEETS = [
    "START_HERE",
    "S1_S5_RUBRIC",
    "STAGE_TAXONOMY",
    "SCOPED_PROFILES",
    "EXCEPTION_RULES",
    "OWNER_DECISIONS",
    "COUNTRY_MODIFIERS_TEMPLATE",
    "GOVERNANCE_TEMPLATE",
    "DATA_DICTIONARY",
]

SUBMISSION_SHEETS = ["S1_S5_RUBRIC", "SCOPE_REFERENCE", "SUBMISSION"]


class WorksheetValidationError(ValueError):
    """Raised when one or more package checks fail."""

    def __init__(self, errors: list[str]):
        self.errors = errors
        super().__init__("\n".join(errors))


def read_csv_exact(
    path: Path, expected_headers: list[str], errors: list[str]
) -> list[dict[str, str]]:
    if not path.exists():
        errors.append(f"Missing required template: {path.name}")
        return []
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != expected_headers:
            errors.append(
                f"{path.name} header must exactly match the canonical schema. "
                f"Expected {expected_headers}; found {reader.fieldnames}."
            )
            return []
        rows: list[dict[str, str]] = []
        for row_number, row in enumerate(reader, start=2):
            if None in row:
                errors.append(f"{path.name}:{row_number} has extra CSV fields.")
                continue
            rows.append({key: value or "" for key, value in row.items()})
        return rows


def validate_taxonomy(
    stages: list[dict[str, str]], errors: list[str]
) -> dict[str, dict[str, str]]:
    stage_by_id: dict[str, dict[str, str]] = {}
    for row_number, row in enumerate(stages, start=2):
        stage_id = row["stage_id"]
        if not stage_id:
            errors.append(f"stages.csv:{row_number} stage_id is required.")
            continue
        if stage_id in stage_by_id:
            errors.append(f"stages.csv has duplicate stage_id {stage_id!r}.")
        stage_by_id[stage_id] = row
        if row["leaf_status"] not in {"leaf", "parent"}:
            errors.append(
                f"stages.csv:{row_number} leaf_status must be leaf or parent."
            )
        if not row["description"].strip():
            errors.append(
                f"stages.csv:{row_number} description must freeze the V1 activity scope."
            )

    for row_number, row in enumerate(stages, start=2):
        parent = row["parent_stage_id"]
        if parent != "not_applicable" and parent not in stage_by_id:
            errors.append(
                f"stages.csv:{row_number} parent_stage_id {parent!r} does not resolve."
            )
        if parent in stage_by_id and stage_by_id[parent]["leaf_status"] != "parent":
            errors.append(
                f"stages.csv:{row_number} parent_stage_id {parent!r} is not a parent row."
            )

    actual_leaves = [row["stage_id"] for row in stages if row["leaf_status"] == "leaf"]
    if actual_leaves != EXPECTED_LEAVES:
        missing = [stage for stage in EXPECTED_LEAVES if stage not in actual_leaves]
        extra = [stage for stage in actual_leaves if stage not in EXPECTED_LEAVES]
        errors.append(
            "stages.csv must contain every frozen leaf exactly once and in canonical "
            f"order. Missing={missing}; extra={extra}; actual_count={len(actual_leaves)}."
        )

    actual_parents = {
        row["stage_id"] for row in stages if row["leaf_status"] == "parent"
    }
    if actual_parents != EXPECTED_FUSION_PARENTS:
        errors.append(
            "stages.csv parent rows must be exactly the eight frozen fusion groups. "
            f"Found {sorted(actual_parents)}."
        )

    quality_scope = stage_by_id.get("quality_assurance", {}).get("description", "")
    if not all(
        phrase in quality_scope
        for phrase in (
            "in-process",
            "end-of-line",
            "ongoing production quality assurance",
            "does not represent a separate regulated product-qualification pathway",
        )
    ):
        errors.append(
            "stages.csv quality_assurance must retain the owner-approved V1 production "
            "quality scope and regulated-qualification exclusion."
        )
    for stage_id in (
        "magnets",
        "heating_and_current_drive",
        "plasma_facing_components",
        "tritium_and_fuel_cycle",
        "blankets",
    ):
        description = stage_by_id.get(stage_id, {}).get("description", "")
        if not all(
            phrase in description
            for phrase in (
                "pilot-relevant",
                "design, engineering, testing, and pre-integration",
                "commercial-fleet rollout",
            )
        ):
            errors.append(
                f"stages.csv {stage_id} must retain the frozen pilot-relevant "
                "subsystem scope and commercial-fleet exclusion."
            )
    if "fabrication after design maturity" not in stage_by_id.get(
        "component_fabrication", {}
    ).get("description", ""):
        errors.append(
            "stages.csv component_fabrication must remain scoped to pilot-relevant "
            "fabrication after design maturity."
        )
    grid_scope = stage_by_id.get("grid_integration", {}).get("description", "")
    if not all(
        phrase in grid_scope
        for phrase in (
            "connection and technical integration for pilot demonstration",
            "Excludes commercial fleet rollout",
        )
    ):
        errors.append(
            "stages.csv grid_integration must retain the pilot-demonstration scope "
            "and commercial-fleet exclusion."
        )
    return stage_by_id


def validate_profiles(
    profiles: list[dict[str, str]],
    stage_by_id: dict[str, dict[str, str]],
    errors: list[str],
) -> dict[str, dict[str, str]]:
    profile_ids = [row["profile_id"] for row in profiles]
    if profile_ids != EXPECTED_PROFILE_IDS:
        errors.append(
            "stage_profiles_template.csv profile_ids must be the stable opaque sequence "
            f"{EXPECTED_PROFILE_IDS[0]} through {EXPECTED_PROFILE_IDS[-1]} in order."
        )
    if len(set(profile_ids)) != len(profile_ids):
        errors.append("stage_profiles_template.csv profile_ids must be unique.")
    for profile_id in profile_ids:
        if not re.fullmatch(r"sp-[0-9]{4}", profile_id):
            errors.append(f"profile_id {profile_id!r} is not a stable opaque package ID.")

    actual_stages = [row["stage_id"] for row in profiles]
    if actual_stages != EXPECTED_LEAVES:
        errors.append(
            "stage_profiles_template.csv must contain one row per frozen leaf in the "
            "same canonical order as stages.csv."
        )

    profile_by_id: dict[str, dict[str, str]] = {}
    for row_number, row in enumerate(profiles, start=2):
        profile_id = row["profile_id"]
        profile_by_id[profile_id] = row
        stage_id = row["stage_id"]
        taxonomy = stage_by_id.get(stage_id)
        if taxonomy is None:
            errors.append(
                f"stage_profiles_template.csv:{row_number} stage_id {stage_id!r} "
                "does not resolve in stages.csv."
            )
            continue
        if taxonomy["leaf_status"] != "leaf":
            errors.append(
                f"stage_profiles_template.csv:{row_number} must reference a leaf stage."
            )
        for field in ("parent_stage_id", "sector"):
            if row[field] != taxonomy[field]:
                errors.append(
                    f"stage_profiles_template.csv:{row_number} {field} must exactly "
                    f"match stages.csv for {stage_id}."
                )
        expected_pathway = EXPECTED_PATHWAYS.get(stage_id)
        if row["pathway_id"] != expected_pathway:
            errors.append(
                f"stage_profiles_template.csv:{row_number} pathway_id must be "
                f"{expected_pathway!r} for {stage_id}."
            )
        if not row["application_context"]:
            errors.append(
                f"stage_profiles_template.csv:{row_number} application_context must "
                "carry the frozen anchor scope."
            )
        expected_lifecycle = EXPECTED_LIFECYCLE_PHASES.get(profile_id)
        if row["lifecycle_phase"] != expected_lifecycle:
            errors.append(
                f"stage_profiles_template.csv:{row_number} lifecycle_phase must be the "
                f"owner-approved primary V1 context {expected_lifecycle!r}."
            )
        if row["critical_path_role"] != "not_assessed":
            errors.append(
                f"stage_profiles_template.csv:{row_number} critical_path_role must be "
                "not_assessed in the blank package."
            )
        for field in PROFILE_HEADERS:
            if field not in SCOPE_FIELDS and row[field]:
                errors.append(
                    f"stage_profiles_template.csv:{row_number} {field} must remain blank "
                    "because only scope/taxonomy fields may be populated."
                )
    return profile_by_id


def validate_review_templates(
    fable_rows: list[dict[str, str]],
    blind_rows: list[dict[str, str]],
    profile_ids: list[str],
    errors: list[str],
) -> None:
    fable_order = [row["profile_id"] for row in fable_rows]
    blind_order = [row["profile_id"] for row in blind_rows]
    if fable_order != profile_ids or blind_order != profile_ids:
        errors.append(
            "Fable and blind CSV templates must contain the identical profile set and "
            "ordering as stage_profiles_template.csv."
        )
    if fable_order != blind_order:
        errors.append("Fable and blind CSV profile sets/order differ.")

    for row_number, row in enumerate(fable_rows, start=2):
        for field, expected in FABLE_IDENTITY.items():
            if row[field] != expected:
                errors.append(
                    f"profile_coding_reviews_fable_template.csv:{row_number} {field} "
                    f"must be {expected!r}."
                )
        for field in REVIEW_HEADERS:
            if field not in {"profile_id", *FABLE_IDENTITY} and row[field]:
                errors.append(
                    f"profile_coding_reviews_fable_template.csv:{row_number} {field} "
                    "must remain blank."
                )

    for row_number, row in enumerate(blind_rows, start=2):
        if row["coder_role"] != "independent_coder":
            errors.append(
                f"profile_coding_reviews_blind_template.csv:{row_number} coder_role "
                "must be the protected provenance value 'independent_coder'."
            )
        for field in REVIEW_HEADERS:
            if field not in {"profile_id", "coder_role"} and row[field]:
                errors.append(
                    f"profile_coding_reviews_blind_template.csv:{row_number} {field} "
                    "must remain blank for the independent task to complete."
                )
        joined = " ".join(row.values()).lower()
        if any(text in joined for text in FORBIDDEN_BLIND_TEXT):
            errors.append(
                f"profile_coding_reviews_blind_template.csv:{row_number} contains "
                "seed-coder identity or hints."
            )


def validate_country_modifier_pathways(
    modifiers: list[dict[str, str]],
    profile_by_id: dict[str, dict[str, str]],
    errors: list[str],
) -> None:
    """Enforce the canonical denormalized profile/pathway referential rule."""
    for row_number, row in enumerate(modifiers, start=2):
        if not any(row.values()):
            continue
        profile_id = row["profile_id"]
        profile = profile_by_id.get(profile_id)
        if profile is None:
            errors.append(
                f"country_stage_modifiers_template.csv:{row_number} profile_id "
                f"{profile_id!r} does not resolve."
            )
            continue
        if row["pathway_id"] != profile["pathway_id"]:
            errors.append(
                f"country_stage_modifiers_template.csv:{row_number} pathway_id "
                f"{row['pathway_id']!r} does not match referenced profile_id "
                f"{profile_id!r} pathway_id {profile['pathway_id']!r}."
            )


def validate_blank_rows(
    filename: str, rows: list[dict[str, str]], errors: list[str]
) -> None:
    populated = [index for index, row in enumerate(rows, start=2) if any(row.values())]
    if populated:
        errors.append(
            f"{filename} must be header-only in the blank package; populated rows: {populated}."
        )


def validate_no_invented_sources(
    named_rows: Iterable[tuple[str, list[dict[str, str]]]], errors: list[str]
) -> None:
    for filename, rows in named_rows:
        for row_number, row in enumerate(rows, start=2):
            for field in row:
                if "source_ids" in field and row[field]:
                    errors.append(
                        f"{filename}:{row_number} {field} must remain blank; source IDs "
                        "may not be invented in this package."
                    )


def validate_no_derived_headers(
    named_headers: Iterable[tuple[str, list[str]]], errors: list[str]
) -> None:
    for filename, headers in named_headers:
        for header in headers:
            lowered = header.lower()
            if any(token in lowered for token in FORBIDDEN_DERIVED_HEADER_TOKENS):
                errors.append(
                    f"{filename} contains prohibited derived field {header!r}; no sector "
                    "scores, averages, weighted composites, or ordinal percentages are allowed."
                )


def validate_human_readable_package(package_dir: Path, errors: list[str]) -> None:
    required_files = {
        "START_HERE.md",
        "RUBRIC.md",
        "OWNER_REVIEW_GUIDE.md",
        "worksheet-build-report.md",
        "unresolved_scope_ambiguities.md",
    }
    contents: dict[str, str] = {}
    for name in sorted(required_files):
        path = package_dir / name
        if not path.exists():
            errors.append(f"Missing required human-readable package file: {name}")
            continue
        contents[name] = path.read_text(encoding="utf-8")

    rubric = contents.get("RUBRIC.md", "")
    normalized_rubric = " ".join(rubric.split())
    for row in RUBRIC_ROWS:
        if row["2_guidance"] not in normalized_rubric:
            errors.append(
                f"RUBRIC.md is missing the exact {row['dimension']} V1 2_guidance."
            )
    for phrase in (
        "operational V1 coding guidance, not a cardinal midpoint",
        "value of 1 or 3",
        "must not be produced by mechanical interpolation",
        "never sum, average, weight, rank, or",
        "percentage-transform S1-S5",
    ):
        if phrase not in normalized_rubric:
            errors.append(f"RUBRIC.md is missing required guidance: {phrase!r}.")

    owner_guide = contents.get("OWNER_REVIEW_GUIDE.md", "")
    normalized_owner_guide = " ".join(owner_guide.split())
    for disposition in ALLOWED_VALUES["owner_disposition"]:
        if f"`{disposition}`" not in owner_guide:
            errors.append(
                f"OWNER_REVIEW_GUIDE.md is missing disposition {disposition!r}."
            )
    if (
        "FICTIONAL EXAMPLE" not in normalized_owner_guide
        or "31 × 5 = 155" not in normalized_owner_guide
    ):
        errors.append(
            "OWNER_REVIEW_GUIDE.md must contain a clearly fictional example and the "
            "155-comparison denominator."
        )

    build_report = contents.get("worksheet-build-report.md", "")
    for phrase in (
        "canonical semantic-and-layout manifest",
        "twice in one temporary runtime",
        "every S1-S5 and C1-C8 field remains blank",
    ):
        if phrase not in build_report:
            errors.append(
                f"worksheet-build-report.md is missing required statement: {phrase!r}."
            )


def validate_exception_schema(path: Path, errors: list[str]) -> None:
    try:
        schema = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        errors.append(f"exception_report.schema.json is unreadable: {exc}")
        return
    try:
        metadata = schema["properties"]["metadata"]["properties"]
        if metadata["comparison_key"]["const"] != ["profile_id", "s_dimension"]:
            errors.append("Exception schema must compare by profile_id and s_dimension.")
        if metadata["expected_comparison_count"].get("const") != 155:
            errors.append("Exception schema must fix the comparison denominator at 155.")
        aggregation = metadata["aggregation_rule"]["const"]
        if not all(token in aggregation for token in ("no_averaging", "midpoint", "consensus")):
            errors.append(
                "Exception schema must explicitly prohibit averaging, midpoint selection, "
                "and forced consensus."
            )
        audit = schema["properties"]["comparison_audit"]
        if audit.get("minItems") != 155 or audit.get("maxItems") != 155:
            errors.append(
                "Exception schema comparison_audit must require exactly 155 rows."
            )
        audit_required = set(schema["$defs"]["comparisonAudit"]["required"])
        expected_audit_fields = {
            "profile_id",
            "stage_id",
            "pathway_id",
            "s_dimension",
            "comparison_status",
            "numeric_difference",
            "fable",
            "independent",
            "semantic_flags",
            "owner_review_required",
            "review_route",
        }
        if audit_required != expected_audit_fields:
            errors.append(
                "Exception schema comparison audit is missing required provenance or "
                "routing fields."
            )
        statuses = schema["$defs"]["comparisonStatus"]["enum"]
        expected_statuses = {
            "exact_agreement",
            "one_point_difference",
            "difference_ge_2",
            "missing_fable",
            "missing_independent",
            "missing_both",
        }
        if set(statuses) != expected_statuses:
            errors.append(
                "Exception schema comparison statuses must contain exactly the six "
                "required audit outcomes."
            )
        summary = schema["properties"]["summary_counts"]
        if set(summary["required"]) != {*expected_statuses, "total_comparisons"}:
            errors.append(
                "Exception schema summary_counts must require all six statuses and a total."
            )
        if summary["properties"]["total_comparisons"].get("const") != 155:
            errors.append("Exception schema summary total must equal 155.")
        flags = schema["$defs"]["semanticFlags"]["items"]["enum"]
        expected_flags = {
            "absolute_difference_ge_2",
            "extreme_0_vs_4",
            "potential_band_or_critical_path_change",
            "contradictory_rationales",
            "missing_or_incompatible_source_support",
            "missing_score",
            "low_confidence_load_bearing_stage",
            "scope_pathway_application_or_lifecycle_ambiguity",
        }
        if set(flags) != expected_flags:
            errors.append("Exception schema does not contain exactly the required flag set.")
        dispositions = schema["$defs"]["exception"]["properties"][
            "owner_disposition"
        ]["enum"]
        if set(dispositions) != {*ALLOWED_VALUES["owner_disposition"], None}:
            errors.append("Exception schema owner dispositions do not match the method task.")
        routine = schema["properties"]["routine_domain_reviews"]
        if routine.get("minItems") != 18 or routine.get("maxItems") != 18:
            errors.append("Exception schema must require all 18 fusion profiles for review.")
        routine_definition = schema["$defs"]["routineDomainReview"]["properties"]
        if routine_definition["owner_decision_required"].get("const") is not False:
            errors.append(
                "Routine fusion domain review must be distinct from an owner decision."
            )
        generation_rules = schema["x-generation-rules"]
        for required_rule in (
            "join",
            "status_and_difference",
            "missing_scores",
            "summary",
            "owner_routing",
            "fusion_routing",
        ):
            if not generation_rules.get(required_rule):
                errors.append(
                    f"Exception schema is missing generation rule {required_rule!r}."
                )
        owner_routing = generation_rules["owner_routing"]
        if not all(
            phrase in owner_routing
            for phrase in (
                "Exact agreements and one-point differences",
                "do not reach owner review by default",
                "only when another semantic flag applies",
            )
        ):
            errors.append(
                "Exception schema must keep exact/one-point rows in the audit and route "
                "one-point differences only when another semantic flag applies."
            )
    except (KeyError, TypeError) as exc:
        errors.append(f"Exception schema is missing required contract elements: {exc}")


def normalized(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def sheet_table(
    workbook_path: Path,
    sheet_name: str,
    expected_headers: list[str],
    errors: list[str],
) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=False)
    if sheet_name not in workbook.sheetnames:
        errors.append(f"{workbook_path.name} is missing sheet {sheet_name}.")
        return []
    sheet = workbook[sheet_name]
    actual_headers = [normalized(sheet.cell(3, column).value) for column in range(1, len(expected_headers) + 1)]
    if actual_headers != expected_headers:
        errors.append(
            f"{workbook_path.name}:{sheet_name} header mismatch. Expected "
            f"{expected_headers}; found {actual_headers}."
        )
        return []
    rows: list[dict[str, str]] = []
    for row_number in range(4, sheet.max_row + 1):
        values = [normalized(sheet.cell(row_number, column).value) for column in range(1, len(expected_headers) + 1)]
        if any(values):
            rows.append(dict(zip(expected_headers, values, strict=True)))
    return rows


def validation_column(validation_range: str) -> int | None:
    first_range = validation_range.split()[0]
    first_cell = first_range.split(":")[0].replace("$", "")
    match = re.match(r"([A-Z]+)[0-9]+", first_cell)
    return column_index_from_string(match.group(1)) if match else None


def workbook_dropdowns(workbook_path: Path, sheet_name: str) -> dict[str, list[str]]:
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[sheet_name]
    headers = {
        column: normalized(sheet.cell(3, column).value)
        for column in range(1, sheet.max_column + 1)
    }
    found: dict[str, list[str]] = {}
    for validation in sheet.data_validations.dataValidation:
        if validation.type != "list":
            continue
        column = validation_column(str(validation.sqref))
        if column is None or not validation.formula1:
            continue
        formula = validation.formula1
        values = formula[1:-1].split(",") if formula.startswith('"') else [formula]
        found[headers.get(column, "")] = values
    return found


def validate_dropdowns(
    workbook_path: Path,
    sheet_name: str,
    expected: dict[str, list[str]],
    errors: list[str],
) -> None:
    actual = workbook_dropdowns(workbook_path, sheet_name)
    if actual != expected:
        errors.append(
            f"{workbook_path.name}:{sheet_name} allowed-value dropdowns must match the "
            f"canonical method. Expected {expected}; found {actual}."
        )


def validate_ordinal_dropdowns(
    workbook_path: Path,
    sheet_name: str,
    fields: set[str],
    errors: list[str],
) -> None:
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[sheet_name]
    headers = {
        column: normalized(sheet.cell(3, column).value)
        for column in range(1, sheet.max_column + 1)
    }
    actual: set[str] = set()
    for validation in sheet.data_validations.dataValidation:
        if validation.type != "whole":
            continue
        if str(validation.formula1) != "0" or str(validation.formula2) != "4":
            errors.append(
                f"{workbook_path.name}:{sheet_name} has a noncanonical ordinal range."
            )
        column = validation_column(str(validation.sqref))
        if column:
            actual.add(headers.get(column, ""))
    if actual != fields:
        errors.append(
            f"{workbook_path.name}:{sheet_name} must validate exactly {sorted(fields)} "
            f"as whole-number 0-4 fields; found {sorted(actual)}."
        )


def validate_workbook_structure(
    workbook_path: Path,
    expected_sheets: list[str],
    errors: list[str],
) -> None:
    if not workbook_path.exists():
        errors.append(f"Missing required workbook: {workbook_path.name}")
        return
    workbook = load_workbook(workbook_path, data_only=False)
    if workbook.sheetnames != expected_sheets:
        errors.append(
            f"{workbook_path.name} sheet order/names mismatch. Expected "
            f"{expected_sheets}; found {workbook.sheetnames}."
        )
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    errors.append(
                        f"{workbook_path.name}:{sheet.title}!{cell.coordinate} contains a "
                        "formula; this blank package must not calculate S/C aggregates."
                    )
        if sheet.title != "START_HERE":
            if not sheet.freeze_panes:
                errors.append(f"{workbook_path.name}:{sheet.title} must freeze headers.")
            if not sheet.auto_filter.ref:
                errors.append(f"{workbook_path.name}:{sheet.title} must enable filters.")


def compare_rows(
    label: str,
    workbook_rows: list[dict[str, str]],
    csv_rows: list[dict[str, str]],
    fields: list[str],
    errors: list[str],
) -> None:
    workbook_view = [{field: row.get(field, "") for field in fields} for row in workbook_rows]
    csv_view = [{field: row.get(field, "") for field in fields} for row in csv_rows]
    if workbook_view != csv_view:
        errors.append(f"{label} does not match its CSV template.")


def validate_workbooks(
    package_dir: Path,
    stages: list[dict[str, str]],
    profiles: list[dict[str, str]],
    fable_rows: list[dict[str, str]],
    blind_rows: list[dict[str, str]],
    errors: list[str],
) -> None:
    reference_path = package_dir / REFERENCE_WORKBOOK
    fable_path = package_dir / FABLE_WORKBOOK
    blind_path = package_dir / BLIND_WORKBOOK

    validate_workbook_structure(reference_path, REFERENCE_SHEETS, errors)
    validate_workbook_structure(fable_path, SUBMISSION_SHEETS, errors)
    validate_workbook_structure(blind_path, SUBMISSION_SHEETS, errors)
    if not all(path.exists() for path in (reference_path, fable_path, blind_path)):
        return

    taxonomy_rows = sheet_table(
        reference_path, "STAGE_TAXONOMY", STAGE_HEADERS, errors
    )
    scoped_rows = sheet_table(
        reference_path, "SCOPED_PROFILES", PROFILE_HEADERS, errors
    )
    owner_rows = sheet_table(reference_path, "OWNER_DECISIONS", OWNER_HEADERS, errors)
    country_rows = sheet_table(
        reference_path, "COUNTRY_MODIFIERS_TEMPLATE", COUNTRY_HEADERS, errors
    )
    governance_rows = sheet_table(
        reference_path, "GOVERNANCE_TEMPLATE", GOVERNANCE_HEADERS, errors
    )
    compare_rows("Reference workbook taxonomy", taxonomy_rows, stages, STAGE_HEADERS, errors)
    compare_rows(
        "Reference workbook scoped profiles",
        scoped_rows,
        profiles,
        PROFILE_HEADERS,
        errors,
    )
    if owner_rows or country_rows or governance_rows:
        errors.append(
            "Reference workbook owner, country-modifier, and governance templates must "
            "contain no populated rows."
        )

    fable_rubric = sheet_table(
        fable_path, "S1_S5_RUBRIC", RUBRIC_HEADERS, errors
    )
    blind_rubric = sheet_table(
        blind_path, "S1_S5_RUBRIC", RUBRIC_HEADERS, errors
    )
    reference_rubric = sheet_table(
        reference_path, "S1_S5_RUBRIC", RUBRIC_HEADERS, errors
    )
    compare_rows(
        "Reference workbook S1-S5 rubric",
        reference_rubric,
        RUBRIC_ROWS,
        RUBRIC_HEADERS,
        errors,
    )
    compare_rows(
        "Fable workbook S1-S5 rubric",
        fable_rubric,
        RUBRIC_ROWS,
        RUBRIC_HEADERS,
        errors,
    )
    compare_rows(
        "Blind workbook S1-S5 rubric",
        blind_rubric,
        RUBRIC_ROWS,
        RUBRIC_HEADERS,
        errors,
    )

    stage_by_id = {row["stage_id"]: row for row in stages}
    expected_scope_rows = [
        {
            "profile_id": profile["profile_id"],
            "stage_id": profile["stage_id"],
            "parent_stage_id": profile["parent_stage_id"],
            "description": stage_by_id.get(profile["stage_id"], {}).get(
                "description", ""
            ),
            "pathway_id": profile["pathway_id"],
            "application_context": profile["application_context"],
            "lifecycle_phase": profile["lifecycle_phase"],
            "critical_path_role": profile["critical_path_role"],
        }
        for profile in profiles
    ]
    fable_scope = sheet_table(
        fable_path, "SCOPE_REFERENCE", SCOPE_REFERENCE_HEADERS, errors
    )
    blind_scope = sheet_table(
        blind_path, "SCOPE_REFERENCE", SCOPE_REFERENCE_HEADERS, errors
    )
    compare_rows(
        "Fable workbook common scope reference",
        fable_scope,
        expected_scope_rows,
        SCOPE_REFERENCE_HEADERS,
        errors,
    )
    compare_rows(
        "Blind workbook common scope reference",
        blind_scope,
        expected_scope_rows,
        SCOPE_REFERENCE_HEADERS,
        errors,
    )
    if fable_scope != blind_scope:
        errors.append(
            "Fable and blind workbooks must contain identical SCOPE_REFERENCE rows."
        )

    fable_submission = sheet_table(
        fable_path, "SUBMISSION", SUBMISSION_HEADERS, errors
    )
    blind_submission = sheet_table(
        blind_path, "SUBMISSION", SUBMISSION_HEADERS, errors
    )
    compare_rows(
        "Fable workbook canonical submission fields",
        fable_submission,
        fable_rows,
        REVIEW_HEADERS,
        errors,
    )
    compare_rows(
        "Blind workbook canonical submission fields",
        blind_submission,
        blind_rows,
        REVIEW_HEADERS,
        errors,
    )
    reference_fields = [
        "profile_id",
        "stage_id",
        "parent_stage_id",
        "sector",
        "workflow",
        "pathway_id",
        "application_context",
        "lifecycle_phase",
        "critical_path_role",
    ]
    compare_rows(
        "Fable workbook scope references",
        fable_submission,
        profiles,
        reference_fields,
        errors,
    )
    compare_rows(
        "Blind workbook scope references",
        blind_submission,
        profiles,
        reference_fields,
        errors,
    )
    fable_order = [row["profile_id"] for row in fable_submission]
    blind_order = [row["profile_id"] for row in blind_submission]
    if fable_order != blind_order or fable_order != EXPECTED_PROFILE_IDS:
        errors.append("Fable and blind workbooks must have identical profile sets/order.")

    for workbook_path, rows in (
        (reference_path, scoped_rows),
        (fable_path, fable_submission),
        (blind_path, blind_submission),
    ):
        for row_number, row in enumerate(rows, start=4):
            for field in S_FIELDS:
                if row.get(field, ""):
                    errors.append(
                        f"{workbook_path.name}:row {row_number} {field} must remain blank."
                    )
            if row.get("coding_confidence", ""):
                errors.append(
                    f"{workbook_path.name}:row {row_number} coding_confidence must be blank."
                )
            for field in row:
                if "source_ids" in field and row[field]:
                    errors.append(
                        f"{workbook_path.name}:row {row_number} {field} must remain blank."
                    )
    for row_number, row in enumerate(country_rows, start=4):
        for field in C_FIELDS:
            if row.get(field, ""):
                errors.append(
                    f"{reference_path.name}:row {row_number} {field} must remain blank."
                )

    blind_workbook = load_workbook(blind_path, data_only=False)
    blind_text = " ".join(
        normalized(cell.value)
        for sheet in blind_workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ).lower()
    blind_metadata = " ".join(
        normalized(value)
        for value in (
            blind_workbook.properties.creator,
            blind_workbook.properties.title,
            blind_workbook.properties.subject,
            blind_workbook.properties.description,
        )
    ).lower()
    if any(text in blind_text or text in blind_metadata for text in FORBIDDEN_BLIND_TEXT):
        errors.append("Blind workbook contains seed-coder identity, values, or hints.")

    blind_submission_sheet = blind_workbook["SUBMISSION"]
    blind_headers = {
        normalized(blind_submission_sheet.cell(3, column).value): column
        for column in range(1, blind_submission_sheet.max_column + 1)
    }
    if not blind_submission_sheet.protection.sheet:
        errors.append("Blind workbook SUBMISSION sheet must enable cell protection.")
    for row_number in range(4, 35):
        role_cell = blind_submission_sheet.cell(
            row_number, blind_headers["coder_role"]
        )
        if role_cell.value != "independent_coder" or not role_cell.protection.locked:
            errors.append(
                f"Blind workbook SUBMISSION row {row_number} must prefill and protect "
                "coder_role=independent_coder."
            )
        for field in ("coder_name", "coder_model"):
            cell = blind_submission_sheet.cell(row_number, blind_headers[field])
            if cell.value not in (None, "") or cell.protection.locked:
                errors.append(
                    f"Blind workbook SUBMISSION row {row_number} {field} must be blank "
                    "and editable by the independent task."
                )

    validate_dropdowns(
        reference_path,
        "STAGE_TAXONOMY",
        {"leaf_status": ALLOWED_VALUES["leaf_status"]},
        errors,
    )
    validate_dropdowns(
        reference_path,
        "SCOPED_PROFILES",
        {
            "lifecycle_phase": ALLOWED_VALUES["lifecycle_phase"],
            "critical_path_role": ALLOWED_VALUES["critical_path_role"],
            "coding_confidence": ALLOWED_VALUES["coding_confidence"],
            "evidence_basis": ALLOWED_VALUES["evidence_basis"],
            "coding_status": ALLOWED_VALUES["coding_status"],
            "review_status": ALLOWED_VALUES["review_status"],
        },
        errors,
    )
    validate_dropdowns(
        reference_path,
        "OWNER_DECISIONS",
        {
            "owner_disposition": ALLOWED_VALUES["owner_disposition"],
            "S_dimension": ALLOWED_VALUES["S_dimension"],
        },
        errors,
    )
    validate_dropdowns(
        reference_path,
        "COUNTRY_MODIFIERS_TEMPLATE",
        {
            "actor_scope": ALLOWED_VALUES["actor_scope"],
            "stage_applicability": ALLOWED_VALUES["stage_applicability"],
            "binding_status": ALLOWED_VALUES["binding_status"],
            "evidence_basis": ALLOWED_VALUES["evidence_basis"],
            "coding_status": ALLOWED_VALUES["coding_status"],
            "review_status": ALLOWED_VALUES["review_status"],
        },
        errors,
    )
    validate_dropdowns(
        reference_path,
        "GOVERNANCE_TEMPLATE",
        {
            "actor_scope": ALLOWED_VALUES["actor_scope"],
            "evidence_basis": ALLOWED_VALUES["evidence_basis"],
            "coding_status": ALLOWED_VALUES["coding_status"],
            "review_status": ALLOWED_VALUES["review_status"],
        },
        errors,
    )
    for path in (fable_path, blind_path):
        validate_dropdowns(
            path,
            "SUBMISSION",
            {
                "coder_type": ALLOWED_VALUES["coder_type"],
                "coding_confidence": ALLOWED_VALUES["coding_confidence"],
                "submission_status": ALLOWED_VALUES["submission_status"],
            },
            errors,
        )
        validate_ordinal_dropdowns(path, "SUBMISSION", set(S_FIELDS), errors)
    validate_ordinal_dropdowns(
        reference_path, "SCOPED_PROFILES", set(S_FIELDS), errors
    )


def manifest_value(value: object) -> object:
    """Return a JSON-serializable scalar without platform-specific repr output."""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.hex()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def color_manifest(color: object | None) -> dict[str, object] | None:
    if color is None:
        return None
    raw = getattr(color, "__dict__", {})
    return {
        "type": raw.get("type"),
        "rgb": raw.get("rgb"),
        "indexed": raw.get("indexed"),
        "auto": raw.get("auto"),
        "theme": raw.get("theme"),
        "tint": raw.get("tint"),
    }


def side_manifest(side: object | None) -> dict[str, object] | None:
    if side is None:
        return None
    return {
        "style": getattr(side, "style", None),
        "color": color_manifest(getattr(side, "color", None)),
    }


def font_manifest(font: object) -> dict[str, object]:
    return {
        "name": getattr(font, "name", None),
        "size": getattr(font, "sz", None),
        "bold": getattr(font, "b", None),
        "italic": getattr(font, "i", None),
        "underline": getattr(font, "u", None),
        "strike": getattr(font, "strike", None),
        "color": color_manifest(getattr(font, "color", None)),
        "vert_align": getattr(font, "vertAlign", None),
        "charset": getattr(font, "charset", None),
        "family": getattr(font, "family", None),
        "scheme": getattr(font, "scheme", None),
        "outline": getattr(font, "outline", None),
        "shadow": getattr(font, "shadow", None),
        "condense": getattr(font, "condense", None),
        "extend": getattr(font, "extend", None),
    }


def fill_manifest(fill: object) -> dict[str, object]:
    stops = []
    for stop in getattr(fill, "stop", ()) or ():
        stops.append(
            {
                "position": getattr(stop, "position", None),
                "color": color_manifest(getattr(stop, "color", None)),
            }
        )
    return {
        "kind": type(fill).__name__,
        "fill_type": getattr(fill, "fill_type", None),
        "pattern_type": getattr(fill, "patternType", None),
        "fg_color": color_manifest(getattr(fill, "fgColor", None)),
        "bg_color": color_manifest(getattr(fill, "bgColor", None)),
        "degree": getattr(fill, "degree", None),
        "left": getattr(fill, "left", None),
        "right": getattr(fill, "right", None),
        "top": getattr(fill, "top", None),
        "bottom": getattr(fill, "bottom", None),
        "stops": sorted(stops, key=lambda item: json.dumps(item, sort_keys=True)),
    }


def border_manifest(border: object) -> dict[str, object]:
    return {
        "left": side_manifest(getattr(border, "left", None)),
        "right": side_manifest(getattr(border, "right", None)),
        "top": side_manifest(getattr(border, "top", None)),
        "bottom": side_manifest(getattr(border, "bottom", None)),
        "diagonal": side_manifest(getattr(border, "diagonal", None)),
        "vertical": side_manifest(getattr(border, "vertical", None)),
        "horizontal": side_manifest(getattr(border, "horizontal", None)),
        "diagonal_up": getattr(border, "diagonalUp", None),
        "diagonal_down": getattr(border, "diagonalDown", None),
        "outline": getattr(border, "outline", None),
    }


def alignment_manifest(alignment: object) -> dict[str, object]:
    return {
        "horizontal": getattr(alignment, "horizontal", None),
        "vertical": getattr(alignment, "vertical", None),
        "text_rotation": getattr(alignment, "textRotation", None),
        "wrap_text": getattr(alignment, "wrapText", None),
        "shrink_to_fit": getattr(alignment, "shrinkToFit", None),
        "indent": getattr(alignment, "indent", None),
        "relative_indent": getattr(alignment, "relativeIndent", None),
        "justify_last_line": getattr(alignment, "justifyLastLine", None),
        "reading_order": getattr(alignment, "readingOrder", None),
    }


def protection_manifest(protection: object) -> dict[str, object]:
    return {
        "locked": getattr(protection, "locked", None),
        "hidden": getattr(protection, "hidden", None),
    }


def comment_manifest(comment: object | None) -> dict[str, object] | None:
    if comment is None:
        return None
    return {
        "text": getattr(comment, "text", None),
        "author": getattr(comment, "author", None),
        "width": getattr(comment, "width", None),
        "height": getattr(comment, "height", None),
    }


def header_footer_manifest(section: object) -> dict[str, object]:
    return {
        "left": getattr(getattr(section, "left", None), "text", None),
        "center": getattr(getattr(section, "center", None), "text", None),
        "right": getattr(getattr(section, "right", None), "text", None),
    }


def workbook_manifest(workbook_path: Path) -> dict[str, Any]:
    """Build a canonical semantic-and-layout manifest for an XLSX workbook."""
    workbook = load_workbook(workbook_path, data_only=False)
    manifest: dict[str, Any] = {
        "properties": {
            "creator": workbook.properties.creator,
            "title": workbook.properties.title,
            "subject": workbook.properties.subject,
            "description": workbook.properties.description,
        },
        "sheet_names": list(workbook.sheetnames),
        "sheets": [],
    }
    for sheet in workbook.worksheets:
        cells: list[dict[str, object]] = []
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        ):
            for cell in row:
                if cell.value is None and cell.comment is None and not cell.has_style:
                    continue
                cells.append(
                    {
                        "coordinate": cell.coordinate,
                        "value": manifest_value(cell.value),
                        "data_type": cell.data_type,
                        "number_format": cell.number_format,
                        "comment": comment_manifest(cell.comment),
                        "font": font_manifest(cell.font),
                        "fill": fill_manifest(cell.fill),
                        "border": border_manifest(cell.border),
                        "alignment": alignment_manifest(cell.alignment),
                        "protection": protection_manifest(cell.protection),
                    }
                )

        row_dimensions = [
            {
                "index": index,
                "height": dimension.height,
                "hidden": dimension.hidden,
                "outline_level": dimension.outlineLevel,
                "collapsed": dimension.collapsed,
                "thick_top": getattr(dimension, "thickTop", None),
                "thick_bottom": getattr(dimension, "thickBot", None),
            }
            for index, dimension in sorted(sheet.row_dimensions.items())
        ]
        column_dimensions = [
            {
                "index": index,
                "width": dimension.width,
                "hidden": dimension.hidden,
                "outline_level": dimension.outlineLevel,
                "collapsed": dimension.collapsed,
                "best_fit": dimension.bestFit,
                "min": dimension.min,
                "max": dimension.max,
            }
            for index, dimension in sorted(
                sheet.column_dimensions.items(),
                key=lambda item: column_index_from_string(item[0]),
            )
        ]
        validations: list[dict[str, object]] = []
        for validation in sheet.data_validations.dataValidation:
            target_ranges = sorted(str(cell_range) for cell_range in validation.ranges)
            validations.append(
                {
                    "type": validation.type,
                    "operator": validation.operator,
                    "formula1": manifest_value(validation.formula1),
                    "formula2": manifest_value(validation.formula2),
                    "allow_blank": validation.allow_blank,
                    "show_error_message": validation.showErrorMessage,
                    "show_input_message": validation.showInputMessage,
                    "error": validation.error,
                    "error_title": validation.errorTitle,
                    "prompt": validation.prompt,
                    "prompt_title": validation.promptTitle,
                    "target_ranges": target_ranges,
                }
            )
        validations.sort(key=lambda item: json.dumps(item, sort_keys=True))

        freeze_panes = sheet.freeze_panes
        if hasattr(freeze_panes, "coordinate"):
            freeze_panes = freeze_panes.coordinate
        page_setup = sheet.page_setup
        page_margins = sheet.page_margins
        print_options = sheet.print_options
        page_setup_properties = sheet.sheet_properties.pageSetUpPr
        protection = sheet.protection
        manifest["sheets"].append(
            {
                "name": sheet.title,
                "visibility": sheet.sheet_state,
                "show_gridlines": sheet.sheet_view.showGridLines,
                "used_range": sheet.calculate_dimension(),
                "cells": cells,
                "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
                "row_dimensions": row_dimensions,
                "column_dimensions": column_dimensions,
                "freeze_panes": freeze_panes,
                "auto_filter": sheet.auto_filter.ref,
                "data_validations": validations,
                "sheet_protection": {
                    "sheet": protection.sheet,
                    "objects": protection.objects,
                    "scenarios": protection.scenarios,
                    "format_cells": protection.formatCells,
                    "format_columns": protection.formatColumns,
                    "format_rows": protection.formatRows,
                    "insert_columns": protection.insertColumns,
                    "insert_rows": protection.insertRows,
                    "delete_columns": protection.deleteColumns,
                    "delete_rows": protection.deleteRows,
                    "select_locked_cells": protection.selectLockedCells,
                    "select_unlocked_cells": protection.selectUnlockedCells,
                    "auto_filter": protection.autoFilter,
                    "sort": protection.sort,
                },
                "print": {
                    "orientation": page_setup.orientation,
                    "paper_size": page_setup.paperSize,
                    "scale": page_setup.scale,
                    "fit_to_width": page_setup.fitToWidth,
                    "fit_to_height": page_setup.fitToHeight,
                    "first_page_number": page_setup.firstPageNumber,
                    "use_first_page_number": page_setup.useFirstPageNumber,
                    "black_and_white": page_setup.blackAndWhite,
                    "draft": page_setup.draft,
                    "cell_comments": page_setup.cellComments,
                    "errors": page_setup.errors,
                    "horizontal_dpi": page_setup.horizontalDpi,
                    "vertical_dpi": page_setup.verticalDpi,
                    "copies": page_setup.copies,
                    "fit_to_page": page_setup_properties.fitToPage,
                    "auto_page_breaks": page_setup_properties.autoPageBreaks,
                    "print_title_rows": sheet.print_title_rows,
                    "print_title_cols": sheet.print_title_cols,
                    "print_area": str(sheet.print_area) if sheet.print_area else None,
                    "horizontal_centered": print_options.horizontalCentered,
                    "vertical_centered": print_options.verticalCentered,
                    "headings": print_options.headings,
                    "grid_lines": print_options.gridLines,
                    "grid_lines_set": print_options.gridLinesSet,
                    "margins": {
                        "left": page_margins.left,
                        "right": page_margins.right,
                        "top": page_margins.top,
                        "bottom": page_margins.bottom,
                        "header": page_margins.header,
                        "footer": page_margins.footer,
                    },
                    "odd_header": header_footer_manifest(sheet.oddHeader),
                    "odd_footer": header_footer_manifest(sheet.oddFooter),
                    "even_header": header_footer_manifest(sheet.evenHeader),
                    "even_footer": header_footer_manifest(sheet.evenFooter),
                    "first_header": header_footer_manifest(sheet.firstHeader),
                    "first_footer": header_footer_manifest(sheet.firstFooter),
                },
            }
        )
    return manifest


def canonical_manifest_json(workbook_path: Path) -> str:
    return json.dumps(
        workbook_manifest(workbook_path),
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def validate_reproducible_workbooks(package_dir: Path, errors: list[str]) -> None:
    input_names = [
        "stages.csv",
        "stage_profiles_template.csv",
        "profile_coding_reviews_fable_template.csv",
        "profile_coding_reviews_blind_template.csv",
        "country_stage_modifiers_template.csv",
        "governance_overlay_template.csv",
        "owner_decisions_template.csv",
    ]
    with tempfile.TemporaryDirectory() as temporary_directory:
        temporary = Path(temporary_directory)
        first_build = temporary / "first_build"
        second_build = temporary / "second_build"
        first_build.mkdir()
        second_build.mkdir()
        for build_directory in (first_build, second_build):
            for name in input_names:
                shutil.copy2(package_dir / name, build_directory / name)
            build_all(build_directory)
        for name in (REFERENCE_WORKBOOK, FABLE_WORKBOOK, BLIND_WORKBOOK):
            committed = package_dir / name
            rebuilt = first_build / name
            if not committed.exists():
                continue
            if canonical_manifest_json(committed) != canonical_manifest_json(rebuilt):
                errors.append(
                    f"{name} semantic-and-layout manifest does not match the rebuilt "
                    "output of scripts/build_structural_profiles_workbooks.py."
                )
            if rebuilt.read_bytes() != (second_build / name).read_bytes():
                errors.append(
                    f"{name} is not byte-identical across two builds in the same "
                    "temporary runtime. Workbook generation is nondeterministic."
                )


def validate(package_dir: Path = DEFAULT_PACKAGE_DIR) -> dict[str, int]:
    package_dir = package_dir.resolve()
    errors: list[str] = []

    stages = read_csv_exact(package_dir / "stages.csv", STAGE_HEADERS, errors)
    profiles = read_csv_exact(
        package_dir / "stage_profiles_template.csv", PROFILE_HEADERS, errors
    )
    fable_rows = read_csv_exact(
        package_dir / "profile_coding_reviews_fable_template.csv",
        REVIEW_HEADERS,
        errors,
    )
    blind_rows = read_csv_exact(
        package_dir / "profile_coding_reviews_blind_template.csv",
        REVIEW_HEADERS,
        errors,
    )
    country_rows = read_csv_exact(
        package_dir / "country_stage_modifiers_template.csv", COUNTRY_HEADERS, errors
    )
    governance_rows = read_csv_exact(
        package_dir / "governance_overlay_template.csv", GOVERNANCE_HEADERS, errors
    )
    owner_rows = read_csv_exact(
        package_dir / "owner_decisions_template.csv", OWNER_HEADERS, errors
    )

    stage_by_id = validate_taxonomy(stages, errors)
    profile_by_id = validate_profiles(profiles, stage_by_id, errors)
    validate_review_templates(
        fable_rows,
        blind_rows,
        [row["profile_id"] for row in profiles],
        errors,
    )
    validate_country_modifier_pathways(country_rows, profile_by_id, errors)
    validate_blank_rows("country_stage_modifiers_template.csv", country_rows, errors)
    validate_blank_rows("governance_overlay_template.csv", governance_rows, errors)
    validate_blank_rows("owner_decisions_template.csv", owner_rows, errors)
    validate_no_invented_sources(
        [
            ("stage_profiles_template.csv", profiles),
            ("profile_coding_reviews_fable_template.csv", fable_rows),
            ("profile_coding_reviews_blind_template.csv", blind_rows),
            ("country_stage_modifiers_template.csv", country_rows),
            ("governance_overlay_template.csv", governance_rows),
            ("owner_decisions_template.csv", owner_rows),
        ],
        errors,
    )
    validate_no_derived_headers(
        [
            ("stage_profiles_template.csv", PROFILE_HEADERS),
            ("profile_coding_reviews_fable_template.csv", REVIEW_HEADERS),
            ("profile_coding_reviews_blind_template.csv", REVIEW_HEADERS),
            ("country_stage_modifiers_template.csv", COUNTRY_HEADERS),
            ("governance_overlay_template.csv", GOVERNANCE_HEADERS),
            ("owner_decisions_template.csv", OWNER_HEADERS),
        ],
        errors,
    )
    validate_human_readable_package(package_dir, errors)
    validate_exception_schema(package_dir / "exception_report.schema.json", errors)
    validate_workbooks(
        package_dir,
        stages,
        profiles,
        fable_rows,
        blind_rows,
        errors,
    )
    if not errors:
        validate_reproducible_workbooks(package_dir, errors)

    if errors:
        raise WorksheetValidationError(errors)
    return {
        "taxonomy_rows": len(stages),
        "leaf_stages": len(EXPECTED_LEAVES),
        "scoped_profiles": len(profiles),
        "fusion_profiles": len(EXPECTED_LEAVES[13:]),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--package-dir",
        type=Path,
        default=DEFAULT_PACKAGE_DIR,
        help="Directory containing the worksheet package.",
    )
    args = parser.parse_args()
    try:
        summary = validate(args.package_dir)
    except WorksheetValidationError as exc:
        print("Structural Profiles worksheet validation failed:")
        for error in exc.errors:
            print(f"- {error}")
        raise SystemExit(1) from exc
    print(
        "Structural Profiles worksheet validation passed: "
        f"{summary['leaf_stages']} leaf stages, "
        f"{summary['scoped_profiles']} scoped profiles, "
        f"{summary['fusion_profiles']} fusion profiles."
    )


if __name__ == "__main__":
    main()
