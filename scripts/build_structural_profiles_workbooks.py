#!/usr/bin/env python3
"""Rebuild the blank Structural Profiles pilot workbooks from package templates."""

from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PACKAGE_DIR = (
    ROOT / "research" / "structural-profiles-pilot" / "worksheet"
)

REFERENCE_WORKBOOK = "structural_profiles_reference_and_owner_review.xlsx"
FABLE_WORKBOOK = "fable_submission_template.xlsx"
BLIND_WORKBOOK = "blind_submission_template.xlsx"

STAGE_HEADERS = [
    "stage_id",
    "parent_stage_id",
    "sector",
    "label",
    "description",
    "display_order",
    "leaf_status",
    "version",
]

PROFILE_HEADERS = [
    "profile_id",
    "stage_id",
    "parent_stage_id",
    "sector",
    "workflow",
    "pathway_id",
    "application_context",
    "lifecycle_phase",
    "critical_path_role",
    "S1",
    "S2",
    "S3",
    "S4",
    "S5",
    "rationale",
    "source_ids",
    "coding_confidence",
    "disagreement_summary",
    "selected_review_ids",
    "evidence_basis",
    "coding_as_of",
    "last_reviewed",
    "revisit_triggers",
    "proposed_by",
    "proposed_model",
    "reviewed_by",
    "independent_review_by",
    "approved_by",
    "coding_status",
    "review_status",
    "version",
    "changelog_note",
]

REVIEW_HEADERS = [
    "review_id",
    "profile_id",
    "coder_type",
    "coder_role",
    "coder_name",
    "coder_model",
    "S1",
    "S2",
    "S3",
    "S4",
    "S5",
    "rationale",
    "source_ids",
    "coding_as_of",
    "submitted_at",
    "submission_status",
    "notes",
]

COUNTRY_HEADERS = [
    "country",
    "subnational_scope",
    "actor_scope",
    "profile_id",
    "pathway_id",
    "stage_applicability",
    "binding_status",
    "C1_accessible_capability",
    "C2_data_access_interoperability",
    "C3_organizational_integration",
    "C4_capital_and_asset_turnover",
    "C5_workforce_integration_skill",
    "C6_governance_procurement_fit",
    "C7_physical_build_test_capacity",
    "C8_ex_ante_commercialization_conditions",
    "source_ids",
    "evidence_basis",
    "confidence",
    "period",
    "coding_as_of",
    "last_reviewed",
    "revisit_triggers",
    "proposed_by",
    "proposed_model",
    "reviewed_by",
    "independent_review_by",
    "approved_by",
    "coding_status",
    "review_status",
    "version",
    "changelog_note",
]

GOVERNANCE_HEADERS = [
    "governance_id",
    "profile_id",
    "country_or_jurisdiction",
    "subnational_scope",
    "actor_scope",
    "benefit_conversion_path",
    "hazard_conversion_path",
    "assurance_burden",
    "assurance_functions",
    "auditability",
    "reversibility",
    "externality_scale",
    "security_sensitivity",
    "concentration_of_control",
    "governance_latency",
    "rationale",
    "source_ids",
    "evidence_basis",
    "confidence",
    "period",
    "last_reviewed",
    "revisit_triggers",
    "proposed_by",
    "proposed_model",
    "reviewed_by",
    "independent_review_by",
    "approved_by",
    "coding_status",
    "review_status",
    "version",
    "changelog_note",
]

OWNER_HEADERS = [
    "decision_id",
    "exception_id",
    "profile_id",
    "stage_id",
    "S_dimension",
    "exception_flags",
    "seed_value",
    "independent_value",
    "seed_rationale",
    "independent_rationale",
    "seed_source_ids",
    "independent_source_ids",
    "owner_disposition",
    "owner_rationale",
    "domain_review_requested",
    "decided_by",
    "decided_at",
    "notes",
]

SUBMISSION_HEADERS = [
    "profile_id",
    "stage_id",
    "parent_stage_id",
    "sector",
    "workflow",
    "pathway_id",
    "application_context",
    "lifecycle_phase",
    "critical_path_role",
    "review_id",
    "coder_type",
    "coder_role",
    "coder_name",
    "coder_model",
    "S1",
    "S2",
    "S3",
    "S4",
    "S5",
    "coding_confidence",
    "rationale",
    "source_ids",
    "coding_as_of",
    "submitted_at",
    "submission_status",
    "notes",
]

SCOPE_REFERENCE_HEADERS = [
    "profile_id",
    "stage_id",
    "parent_stage_id",
    "description",
    "pathway_id",
    "application_context",
    "lifecycle_phase",
    "critical_path_role",
]

RUBRIC_HEADERS = [
    "dimension",
    "name",
    "direction",
    "0_means",
    "2_guidance",
    "4_means",
    "coding_instruction",
]

RUBRIC_ROWS = [
    {
        "dimension": "S1",
        "name": "Information intensity of the scoped stage",
        "direction": "Higher means structurally easier for AI-driven improvement.",
        "0_means": (
            "Progress occurs primarily through physical transformation, biological "
            "processes, construction, or accumulated operating time."
        ),
        "2_guidance": (
            "Stage progress materially combines informational and physical, biological, "
            "or operational work; neither clearly dominates under the stated scope."
        ),
        "4_means": (
            "Progress occurs primarily through information processing, analysis, "
            "design, inference, communication, or software."
        ),
        "coding_instruction": (
            "Code how progress occurs within this stage, not its share of a wider "
            "schedule. Critical-path role is separate. Values 1 and 3 require an "
            "explicit intermediate-judgment rationale; do not interpolate mechanically."
        ),
    },
    {
        "dimension": "S2",
        "name": "Feedback speed",
        "direction": "Higher means structurally easier for AI-driven improvement.",
        "0_means": "Cycles take years to decades.",
        "2_guidance": (
            "The ordinary learn-test-revise cycle is generally measured in weeks to "
            "months, rather than minutes/days or years/decades."
        ),
        "4_means": "Learn-test-revise cycles occur in minutes to days.",
        "coding_instruction": (
            "Classify the ordinary scoped feedback loop, not an ideal demo. Values 1 "
            "and 3 require an explicit intermediate-judgment rationale; do not "
            "interpolate mechanically."
        ),
    },
    {
        "dimension": "S3",
        "name": "Experiment affordability and throughput",
        "direction": "Higher means structurally easier for AI-driven improvement.",
        "0_means": "Attempts are scarce and expensive.",
        "2_guidance": (
            "Attempts require meaningful resources or scarce systems but remain "
            "repeatable; throughput or parallelism is material but constrained."
        ),
        "4_means": "Attempts are near-free and parallel.",
        "coding_instruction": (
            "The rationale must address both marginal cost per attempt and attainable "
            "parallel or serial throughput. Preserve divergence and lower confidence. "
            "Values 1 and 3 require an explicit intermediate-judgment rationale; do "
            "not interpolate mechanically."
        ),
    },
    {
        "dimension": "S4",
        "name": "Physical flexibility",
        "direction": "Higher means structurally easier for AI-driven improvement.",
        "0_means": "Clock time is dominated by physical floors.",
        "2_guidance": (
            "Physical processes create meaningful delay, but no single physical, "
            "biological, construction, curing, healing, or qualification floor "
            "dominates the entire scoped stage."
        ),
        "4_means": (
            "No construction, growth, curing, healing, or qualification floor "
            "dominates elapsed time."
        ),
        "coding_instruction": (
            "Code intrinsic physical time floors at the scoped stage. Values 1 and 3 "
            "require an explicit intermediate-judgment rationale; do not interpolate "
            "mechanically."
        ),
    },
    {
        "dimension": "S5",
        "name": "Intrinsic error tolerance",
        "direction": "Higher means structurally easier for AI-driven improvement.",
        "0_means": "Errors are catastrophic or irreversible.",
        "2_guidance": (
            "Errors create material cost, delay, or review burdens, but are usually "
            "containable and reversible within the stated scope."
        ),
        "4_means": "Errors are cheap, reversible, and low-externality.",
        "coding_instruction": (
            "Classify consequences and reversibility within scope. Values 1 and 3 "
            "require an explicit intermediate-judgment rationale; do not interpolate "
            "mechanically."
        ),
    },
]

EXCEPTION_HEADERS = [
    "rule_id",
    "trigger",
    "owner_review",
    "domain_review",
    "required_handling",
]

EXCEPTION_ROWS = [
    {
        "rule_id": "absolute_difference_ge_2",
        "trigger": "Absolute seed-versus-independent difference is 2 or more on an S-dimension.",
        "owner_review": "Yes",
        "domain_review": "When technically load-bearing",
        "required_handling": "Preserve both values and rationales; do not average.",
    },
    {
        "rule_id": "extreme_0_vs_4",
        "trigger": "One coder assigns 0 and the other assigns 4.",
        "owner_review": "Yes",
        "domain_review": "Usually",
        "required_handling": "Flag even when another disagreement rule also applies.",
    },
    {
        "rule_id": "potential_band_or_critical_path_change",
        "trigger": "The disagreement could change the stage band or critical-path reading.",
        "owner_review": "Yes",
        "domain_review": "When scope expertise is needed",
        "required_handling": "Review the stated scope; do not infer a sector average.",
    },
    {
        "rule_id": "contradictory_rationales",
        "trigger": "The explanations rely on incompatible mechanisms or scope readings.",
        "owner_review": "Yes",
        "domain_review": "As needed",
        "required_handling": "Retain both explanations and identify the contradiction.",
    },
    {
        "rule_id": "missing_or_incompatible_source_support",
        "trigger": "Support is missing or the two source bases do not support the same scope.",
        "owner_review": "Yes",
        "domain_review": "As needed",
        "required_handling": "Route to needs_better_evidence when necessary; never invent a source ID.",
    },
    {
        "rule_id": "missing_score",
        "trigger": "Either submission is missing or has an invalid S value for the comparison.",
        "owner_review": "Yes",
        "domain_review": "As needed",
        "required_handling": "Keep the audit row, mark missing_score, and never silently omit it.",
    },
    {
        "rule_id": "low_confidence_load_bearing_stage",
        "trigger": "A load-bearing stage has low coding confidence.",
        "owner_review": "Yes",
        "domain_review": "Yes",
        "required_handling": "Route for evidence or domain review without manufacturing certainty.",
    },
    {
        "rule_id": "scope_pathway_application_or_lifecycle_ambiguity",
        "trigger": "A scope, pathway, application, or lifecycle field is ambiguous.",
        "owner_review": "Yes",
        "domain_review": "As needed",
        "required_handling": "Resolve scope before selecting a value; blank is allowed.",
    },
    {
        "rule_id": "routine_fusion_domain_review",
        "trigger": "Every frozen fusion profile, whether or not another flag fires.",
        "owner_review": "No, unless another exception applies",
        "domain_review": "Yes",
        "required_handling": (
            "Queue as routine domain review; keep it distinct from an owner decision."
        ),
    },
]

ALLOWED_VALUES = {
    "actor_scope": [
        "national_aggregate",
        "large_firms",
        "smes",
        "public_sector",
        "research_institutions",
        "households",
        "defense_establishment",
        "not_assessable",
    ],
    "assurance_functions": [
        "error_detection",
        "trust_production",
        "delay",
        "experimentation_barrier",
    ],
    "coding_confidence": ["low", "medium", "high"],
    "coding_status": ["proposed", "reviewed", "approved", "disputed"],
    "coder_type": ["human", "model"],
    "critical_path_role": [
        "serial",
        "parallel",
        "conditional",
        "time_floor",
        "not_assessed",
    ],
    "evidence_basis": [
        "observed",
        "model estimate",
        "scenario",
        "official target",
        "company target",
        "expert-coded",
        "historical analogy",
        "hypothesis",
    ],
    "leaf_status": ["leaf", "parent"],
    "lifecycle_phase": [
        "research",
        "development",
        "demonstration",
        "qualification",
        "scale_up",
        "commercial_deployment",
        "operations",
        "diffusion",
    ],
    "owner_disposition": [
        "prefer_seed",
        "prefer_independent",
        "preserve_disagreement",
        "needs_domain_review",
        "needs_better_evidence",
    ],
    "review_status": ["canonical", "reviewed", "staged", "superseded", "rejected"],
    "stage_applicability": ["present", "absent", "conditional", "not_assessable"],
    "binding_status": ["binding", "non_binding", "contested", "not_assessable"],
    "submission_status": ["submitted", "withdrawn", "superseded"],
    "S_dimension": ["S1", "S2", "S3", "S4", "S5"],
}

FIELD_DEFINITIONS = {
    "profile_id": "Stable opaque identifier for one scoped stage profile.",
    "stage_id": "Stable stage identifier from stages.csv.",
    "description": "Frozen V1 statement of activity included and material exclusions.",
    "parent_stage_id": "Parent from stages.csv, or not_applicable for a root stage.",
    "sector": "Frozen anchor-sector label.",
    "workflow": "Readable workflow label for the scoped leaf stage.",
    "pathway_id": "Frozen technical or operational pathway identifier.",
    "application_context": "Frozen V1 anchor scope copied from the canonical method.",
    "lifecycle_phase": (
        "Owner-approved primary V1 coding context; it does not claim that the stage "
        "occurs exclusively in that lifecycle phase."
    ),
    "critical_path_role": "Serial, parallel, conditional, time floor, or not yet assessed.",
    "rationale": "Coder explanation for the proposed classification.",
    "source_ids": "Identifiers for actual supporting sources; never invent them.",
    "coding_confidence": "Low, medium, or high confidence in the scoped coding.",
    "coding_as_of": "Evidence vintage for a proposed coding.",
    "last_reviewed": "Date the row was last substantively reviewed.",
    "revisit_triggers": "Concrete events or new vintages that require row review.",
    "selected_review_ids": "Submission IDs used in the current profile disposition.",
    "disagreement_summary": "Dimension-specific unresolved differences; never a midpoint.",
    "evidence_basis": "What kind of evidence supports the row.",
    "coding_status": "Coder-workflow status, separate from public review status.",
    "review_status": "Global rendering/review gate, separate from coding status.",
    "review_id": "Stable identifier assigned to a completed coder submission.",
    "coder_type": "Human or model.",
    "coder_role": "Documented role of the coder in the review workflow.",
    "coder_name": "Person, provider, or documented reviewer identity.",
    "coder_model": "Model identifier for a model submission.",
    "submission_status": "Record status; it does not grant approval.",
    "country": "Country attached to the country-stage modifier.",
    "subnational_scope": "Covered jurisdiction or a canonical non-national sentinel.",
    "actor_scope": "Actor class to which the evidence applies.",
    "stage_applicability": "Whether the profile is present, absent, conditional, or unassessable.",
    "binding_status": "Whether the stage is binding in the stated country/pathway scope.",
    "period": "Measurement or evidence period for the modifier or governance row.",
    "governance_id": "Stable identifier for one governance-overlay record.",
    "assurance_functions": "One or more assurance functions named in the method.",
    "owner_disposition": "Owner action selected from the five allowed exception outcomes.",
    "S_dimension": "The individual S-dimension involved in an exception.",
    "exception_flags": "One or more machine-generated reasons the row needs attention.",
    "domain_review_requested": "Whether subject-matter review is requested after triage.",
}

PALETTE = {
    "ink": "24313D",
    "header": "35495E",
    "header_text": "FFFFFF",
    "instruction": "DCEAF4",
    "reference": "E8ECEF",
    "editable": "FFF2CC",
    "white": "FFFFFF",
    "hairline": "CFD6DC",
}

THIN_BOTTOM = Border(bottom=Side(style="thin", color=PALETTE["hairline"]))
FIXED_TIMESTAMP = datetime(2026, 8, 30, 0, 0, 0, tzinfo=timezone.utc)


def read_csv(path: Path, expected_headers: list[str]) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != expected_headers:
            raise ValueError(
                f"{path.name} header mismatch. Expected {expected_headers}; "
                f"found {reader.fieldnames}."
            )
        return [{key: value or "" for key, value in row.items()} for row in reader]


def set_properties(workbook: Workbook, title: str) -> None:
    workbook.properties.creator = "AI Conversion Atlas"
    workbook.properties.title = title
    workbook.properties.subject = "Blank Structural Profiles pilot research worksheet"
    workbook.properties.description = (
        "Blank owner-reviewable research classification package; no S/C values assigned."
    )
    workbook.properties.created = FIXED_TIMESTAMP
    workbook.properties.modified = FIXED_TIMESTAMP


def list_validation(values: list[str]) -> DataValidation:
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Choose a value from the canonical list or leave blank."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Use a canonical allowed value. Blank remains allowed."
    validation.promptTitle = "Allowed values"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    return validation


def ordinal_validation() -> DataValidation:
    validation = DataValidation(
        type="whole",
        operator="between",
        formula1="0",
        formula2="4",
        allow_blank=True,
    )
    validation.error = "Enter a whole number from 0 to 4, or leave blank."
    validation.errorTitle = "Invalid ordinal value"
    validation.showErrorMessage = True
    return validation


def preferred_width(header: str, values: list[str]) -> int:
    sample = max([len(header), *(len(str(value)) for value in values[:60])])
    if header in {
        "application_context",
        "description",
        "rationale",
        "notes",
        "required_handling",
        "2_guidance",
    }:
        return 48
    if "source_ids" in header or header.endswith("_id") or header.endswith("_ids"):
        return min(32, max(18, sample + 2))
    if header in {"workflow", "trigger", "coding_instruction", "definition"}:
        return min(44, max(22, sample // 2 + 8))
    return min(28, max(12, sample + 2))


def add_table_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[dict[str, str]],
    note: str,
    *,
    editable_fields: set[str] | None = None,
    reserve_rows: int = 0,
    dropdown_fields: dict[str, list[str]] | None = None,
    ordinal_fields: set[str] | None = None,
) -> None:
    editable_fields = editable_fields or set()
    dropdown_fields = dropdown_fields or {}
    ordinal_fields = ordinal_fields or set()

    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    last_column = get_column_letter(len(headers))
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    note_cell = sheet.cell(1, 1, note)
    note_cell.fill = PatternFill("solid", fgColor=PALETTE["instruction"])
    note_cell.font = Font(name="Aptos", size=10, color=PALETTE["ink"])
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[1].height = 48
    sheet.row_dimensions[2].height = 8

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(3, column, header)
        cell.fill = PatternFill("solid", fgColor=PALETTE["header"])
        cell.font = Font(name="Aptos", size=10, bold=True, color=PALETTE["header_text"])
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        role = "editable later" if header in editable_fields else "generated/reference"
        cell.comment = Comment(f"Cell role: {role}.", "AI Conversion Atlas")
        cell.protection = Protection(locked=True)
    sheet.row_dimensions[3].height = 42

    for row_offset, record in enumerate(rows, start=4):
        for column, header in enumerate(headers, start=1):
            sheet.cell(row_offset, column, record.get(header, ""))

    final_row = max(3 + len(rows), 3 + reserve_rows)
    for row_number in range(4, final_row + 1):
        sheet.row_dimensions[row_number].height = 36
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row_number, column)
            fill = PALETTE["editable"] if header in editable_fields else PALETTE["reference"]
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="Aptos", size=10, color=PALETTE["ink"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BOTTOM
            cell.number_format = "@"
            cell.protection = Protection(locked=header not in editable_fields)

    for column, header in enumerate(headers, start=1):
        values = [record.get(header, "") for record in rows]
        sheet.column_dimensions[get_column_letter(column)].width = preferred_width(
            header, values
        )

    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:{last_column}{final_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 2 if len(headers) > 20 else 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_title_rows = "1:3"
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    sheet.protection.sheet = True
    sheet.protection.autoFilter = False
    sheet.protection.sort = False

    field_to_column = {header: index for index, header in enumerate(headers, start=1)}
    for field, values in dropdown_fields.items():
        if field not in field_to_column:
            continue
        validation = list_validation(values)
        sheet.add_data_validation(validation)
        letter = get_column_letter(field_to_column[field])
        validation.add(f"{letter}4:{letter}{final_row}")
    for field in sorted(ordinal_fields):
        if field not in field_to_column:
            continue
        validation = ordinal_validation()
        sheet.add_data_validation(validation)
        letter = get_column_letter(field_to_column[field])
        validation.add(f"{letter}4:{letter}{final_row}")


def add_start_here_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("START_HERE")
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.column_dimensions["A"].width = 24
    for column in range(2, 10):
        sheet.column_dimensions[get_column_letter(column)].width = 15

    sheet.merge_cells("A1:I1")
    title = sheet["A1"]
    title.value = "Structural Profiles pilot — owner reference and review workbook"
    title.fill = PatternFill("solid", fgColor=PALETTE["header"])
    title.font = Font(name="Aptos Display", size=16, bold=True, color=PALETTE["header_text"])
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34

    blocks = [
        (
            "Purpose",
            "This blank package tests whether two models can apply S1-S5 consistently "
            "to the same 31 scoped leaf stages. Coding means research classification, "
            "not software programming. No S/C value or owner decision exists yet.",
        ),
        (
            "Roles",
            "The completed seed coding was produced in Claude Code using Claude Opus "
            "5. The independent coding was produced by Codex using GPT-5.6 at "
            "extra-high reasoning effort. Fable remains the framework architect, not "
            "the row-level coder. Jinhua reviews routed exceptions only; domain "
            "experts later review all fusion rows and other routed cases.",
        ),
        (
            "Owner workload",
            "Jinhua does not review all 155 S-cells. The later report flags material "
            "numeric disagreement, extreme 0-versus-4 cases, interpretation changes, "
            "contradictory rationales, missing support, low-confidence load-bearing "
            "stages, and scope ambiguity.",
        ),
        (
            "Safe handoff",
            "The original Fable-named template is a historical handoff artifact. For "
            "reconciliation, use the immutable seed_submission_v1.csv and "
            "independent_submission_v1.csv copies pinned to their recorded PR heads.",
        ),
        (
            "Sequence",
            "Separate coding tasks → validate returns → generate exceptions without "
            "averaging → owner reviews exceptions → domain review (all fusion rows) → "
            "separate WP2 implementation/population task.",
        ),
        (
            "Blank means blank",
            "S1-S5, C1-C8, rationales, source IDs, dates, confidence, notes, "
            "country/governance rows, exception results, owner decisions, and "
            "all approval/canonicalization fields must remain blank in this package.",
        ),
        (
            "Cell colors",
            "Gray cells are generated/reference structure. Yellow cells are editable "
            "only in a later named task. Blue rows are instructions. Blank never means "
            "zero, agreement, not applicable, or approval.",
        ),
    ]

    row = 3
    for heading, body in blocks:
        sheet.cell(row, 1, heading)
        sheet.cell(row, 1).font = Font(name="Aptos", size=11, bold=True, color=PALETTE["ink"])
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=PALETTE["instruction"])
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        body_cell = sheet.cell(row, 2, body)
        body_cell.font = Font(name="Aptos", size=10, color=PALETTE["ink"])
        body_cell.alignment = Alignment(wrap_text=True, vertical="top")
        body_cell.fill = PatternFill("solid", fgColor=PALETTE["white"])
        sheet.row_dimensions[row].height = 62
        row += 2


def profile_dropdowns() -> dict[str, list[str]]:
    return {
        "lifecycle_phase": ALLOWED_VALUES["lifecycle_phase"],
        "critical_path_role": ALLOWED_VALUES["critical_path_role"],
        "coding_confidence": ALLOWED_VALUES["coding_confidence"],
        "evidence_basis": ALLOWED_VALUES["evidence_basis"],
        "coding_status": ALLOWED_VALUES["coding_status"],
        "review_status": ALLOWED_VALUES["review_status"],
    }


def add_data_dictionary(workbook: Workbook) -> None:
    headers = [
        "object",
        "field",
        "plain_english_definition",
        "allowed_values",
        "cell_role",
        "blank_package_state",
    ]
    objects = [
        ("stages.csv", STAGE_HEADERS, set()),
        ("stage_profiles_template.csv", PROFILE_HEADERS, set()),
        ("profile_coding_reviews_*_template.csv", REVIEW_HEADERS, set(REVIEW_HEADERS) - {"profile_id"}),
        ("country_stage_modifiers_template.csv", COUNTRY_HEADERS, set(COUNTRY_HEADERS)),
        ("governance_overlay_template.csv", GOVERNANCE_HEADERS, set(GOVERNANCE_HEADERS)),
        ("owner_decisions_template.csv", OWNER_HEADERS, {
            "owner_disposition",
            "owner_rationale",
            "domain_review_requested",
            "decided_by",
            "decided_at",
            "notes",
        }),
        ("submission workbooks only", ["coding_confidence"], {"coding_confidence"}),
    ]
    rows: list[dict[str, str]] = []
    for object_name, fields, editable in objects:
        for field in fields:
            if field in {"S1", "S2", "S3", "S4", "S5"}:
                definition = "Ordinal intrinsic-stage dimension; whole number 0-4 when coded."
                allowed = "0 | 1 | 2 | 3 | 4"
            elif field.startswith("C") and "_" in field:
                definition = (
                    "Ordinal country-stage conversion condition; higher means more enabling."
                )
                allowed = "Canonical method does not yet fix a numeric dropdown range."
            elif field == "subnational_scope":
                definition = FIELD_DEFINITIONS[field]
                allowed = "national | not_applicable | not_assessable | named jurisdiction"
            else:
                definition = FIELD_DEFINITIONS.get(
                    field, field.replace("_", " ").capitalize() + "."
                )
                allowed = " | ".join(ALLOWED_VALUES.get(field, []))
            rows.append(
                {
                    "object": object_name,
                    "field": field,
                    "plain_english_definition": definition,
                    "allowed_values": allowed,
                    "cell_role": "editable later" if field in editable else "generated/reference",
                    "blank_package_state": (
                        "blank" if field not in {
                            "profile_id",
                            "stage_id",
                            "parent_stage_id",
                            "sector",
                            "workflow",
                            "pathway_id",
                            "application_context",
                            "lifecycle_phase",
                            "critical_path_role",
                            "display_order",
                            "leaf_status",
                            "label",
                            "description",
                            "coder_type",
                            "coder_role",
                            "coder_name",
                            "coder_model",
                            "version",
                        } else "reference populated where supported"
                    ),
                }
            )
    add_table_sheet(
        workbook,
        "DATA_DICTIONARY",
        headers,
        rows,
        "Reference only. Allowed-value lists reproduce the canonical method; blank "
        "means the method does not fix a list or the field is intentionally empty.",
    )


def build_reference_workbook(package_dir: Path) -> Path:
    stages = read_csv(package_dir / "stages.csv", STAGE_HEADERS)
    profiles = read_csv(package_dir / "stage_profiles_template.csv", PROFILE_HEADERS)
    owner_rows = read_csv(package_dir / "owner_decisions_template.csv", OWNER_HEADERS)
    country_rows = read_csv(
        package_dir / "country_stage_modifiers_template.csv", COUNTRY_HEADERS
    )
    governance_rows = read_csv(
        package_dir / "governance_overlay_template.csv", GOVERNANCE_HEADERS
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    set_properties(workbook, "Structural Profiles reference and owner review")
    add_start_here_sheet(workbook)
    add_table_sheet(
        workbook,
        "S1_S5_RUBRIC",
        RUBRIC_HEADERS,
        RUBRIC_ROWS,
        "Reference only. Higher is always structurally easier. Keep all five dimensions "
        "separate; never sum, average, weight, rank, or percentage-transform them. "
        "The 2 anchor is operational guidance, not a cardinal midpoint or new method rule.",
    )
    add_table_sheet(
        workbook,
        "STAGE_TAXONOMY",
        STAGE_HEADERS,
        stages,
        "Generated/reference cells (gray) reproduce stages.csv. Do not edit in the workbook.",
        dropdown_fields={"leaf_status": ALLOWED_VALUES["leaf_status"]},
    )
    add_table_sheet(
        workbook,
        "SCOPED_PROFILES",
        PROFILE_HEADERS,
        profiles,
        "Generated/reference view of stage_profiles_template.csv. All judgment, evidence, "
        "review, and approval fields must remain blank in this package.",
        dropdown_fields=profile_dropdowns(),
        ordinal_fields={"S1", "S2", "S3", "S4", "S5"},
    )
    add_table_sheet(
        workbook,
        "EXCEPTION_RULES",
        EXCEPTION_HEADERS,
        EXCEPTION_ROWS,
        "Generated specification. Compare by profile_id and individual S-dimension; no "
        "averaging or forced consensus. Routine fusion review is not automatically an owner decision.",
    )
    add_table_sheet(
        workbook,
        "OWNER_DECISIONS",
        OWNER_HEADERS,
        owner_rows,
        "Generated exception details will be gray; later owner-editable cells are yellow. "
        "No rows exist now because Jinhua reviews only generated exceptions, not all 155 S-cells.",
        editable_fields={
            "owner_disposition",
            "owner_rationale",
            "domain_review_requested",
            "decided_by",
            "decided_at",
            "notes",
        },
        reserve_rows=100,
        dropdown_fields={
            "owner_disposition": ALLOWED_VALUES["owner_disposition"],
            "S_dimension": ALLOWED_VALUES["S_dimension"],
        },
    )
    add_table_sheet(
        workbook,
        "COUNTRY_MODIFIERS_TEMPLATE",
        COUNTRY_HEADERS,
        country_rows,
        "Later researcher-editable cells are yellow. The template is header-only and all "
        "C fields remain blank. pathway_id must match the referenced profile_id exactly.",
        editable_fields=set(COUNTRY_HEADERS),
        reserve_rows=50,
        dropdown_fields={
            "actor_scope": ALLOWED_VALUES["actor_scope"],
            "stage_applicability": ALLOWED_VALUES["stage_applicability"],
            "binding_status": ALLOWED_VALUES["binding_status"],
            "evidence_basis": ALLOWED_VALUES["evidence_basis"],
            "coding_status": ALLOWED_VALUES["coding_status"],
            "review_status": ALLOWED_VALUES["review_status"],
        },
    )
    add_table_sheet(
        workbook,
        "GOVERNANCE_TEMPLATE",
        GOVERNANCE_HEADERS,
        governance_rows,
        "Later researcher-editable cells are yellow. The template is header-only and "
        "separate from intrinsic S and country-stage C fields.",
        editable_fields=set(GOVERNANCE_HEADERS),
        reserve_rows=50,
        dropdown_fields={
            "actor_scope": ALLOWED_VALUES["actor_scope"],
            "evidence_basis": ALLOWED_VALUES["evidence_basis"],
            "coding_status": ALLOWED_VALUES["coding_status"],
            "review_status": ALLOWED_VALUES["review_status"],
        },
    )
    add_data_dictionary(workbook)

    output = package_dir / REFERENCE_WORKBOOK
    workbook.save(output)
    normalize_xlsx(output)
    return output


def build_submission_workbook(
    package_dir: Path,
    *,
    review_template: str,
    output_name: str,
    blind: bool,
) -> Path:
    profiles = read_csv(package_dir / "stage_profiles_template.csv", PROFILE_HEADERS)
    stages = read_csv(package_dir / "stages.csv", STAGE_HEADERS)
    reviews = read_csv(package_dir / review_template, REVIEW_HEADERS)
    profile_by_id = {row["profile_id"]: row for row in profiles}
    stage_by_id = {row["stage_id"]: row for row in stages}
    submission_rows: list[dict[str, str]] = []
    for review in reviews:
        profile = profile_by_id[review["profile_id"]]
        record = {header: "" for header in SUBMISSION_HEADERS}
        for header in SUBMISSION_HEADERS:
            if header in profile:
                record[header] = profile[header]
            if header in review:
                record[header] = review[header]
        record["coding_confidence"] = ""
        submission_rows.append(record)

    workbook = Workbook()
    workbook.remove(workbook.active)
    title = (
        "Blind independent Structural Profiles submission"
        if blind
        else "Fable Structural Profiles seed submission"
    )
    set_properties(workbook, title)
    add_table_sheet(
        workbook,
        "S1_S5_RUBRIC",
        RUBRIC_HEADERS,
        RUBRIC_ROWS,
        "Reference only. Higher is always structurally easier. Keep dimensions separate; "
        "do not sum, average, weight, rank, or percentage-transform them. The 2 anchor "
        "is operational guidance, not a cardinal midpoint or new method rule.",
    )
    scope_rows = [
        {
            "profile_id": profile["profile_id"],
            "stage_id": profile["stage_id"],
            "parent_stage_id": profile["parent_stage_id"],
            "description": stage_by_id[profile["stage_id"]]["description"],
            "pathway_id": profile["pathway_id"],
            "application_context": profile["application_context"],
            "lifecycle_phase": profile["lifecycle_phase"],
            "critical_path_role": profile["critical_path_role"],
        }
        for profile in profiles
    ]
    add_table_sheet(
        workbook,
        "SCOPE_REFERENCE",
        SCOPE_REFERENCE_HEADERS,
        scope_rows,
        "Read-only common scope. Both submission workbooks contain this identical "
        "profile set, order, descriptions, and primary V1 lifecycle context.",
    )
    editable = {
        "review_id",
        "S1",
        "S2",
        "S3",
        "S4",
        "S5",
        "coding_confidence",
        "rationale",
        "source_ids",
        "coding_as_of",
        "submitted_at",
        "submission_status",
        "notes",
    }
    if blind:
        editable |= {"coder_type", "coder_name", "coder_model"}
        note = (
            "Gray cells are frozen scope references. Yellow cells must be completed by "
            "the independent coder. coder_role=independent_coder is protected; the "
            "independent task fills coder_type, coder_name, and coder_model. Do not use "
            "another coder's submission or reasoning."
        )
    else:
        note = (
            "Gray cells are frozen scope references or authoritative coder provenance. "
            "Yellow cells are the later seed submission. S1-S5, rationale, source_ids, "
            "coding_as_of, coding_confidence, and notes are blank now."
        )
    add_table_sheet(
        workbook,
        "SUBMISSION",
        SUBMISSION_HEADERS,
        submission_rows,
        note,
        editable_fields=editable,
        dropdown_fields={
            "coder_type": ALLOWED_VALUES["coder_type"],
            "coding_confidence": ALLOWED_VALUES["coding_confidence"],
            "submission_status": ALLOWED_VALUES["submission_status"],
        },
        ordinal_fields={"S1", "S2", "S3", "S4", "S5"},
    )
    output = package_dir / output_name
    workbook.save(output)
    normalize_xlsx(output)
    return output


def normalize_xlsx(path: Path) -> None:
    """Normalize ZIP member order/timestamps so repeated builds are byte-stable."""
    with ZipFile(path, "r") as source:
        members = {name: source.read(name) for name in source.namelist()}
    core = members.get("docProps/core.xml")
    if core:
        members["docProps/core.xml"] = re.sub(
            rb"(<dcterms:modified[^>]*>).*?(</dcterms:modified>)",
            rb"\g<1>2026-08-30T00:00:00Z\g<2>",
            core,
        )
    temporary = path.with_suffix(".normalized.xlsx")
    with ZipFile(temporary, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
        for name in sorted(members):
            info = ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, members[name])
    temporary.replace(path)


def build_all(package_dir: Path = DEFAULT_PACKAGE_DIR) -> list[Path]:
    package_dir.mkdir(parents=True, exist_ok=True)
    outputs = [
        build_reference_workbook(package_dir),
        build_submission_workbook(
            package_dir,
            review_template="profile_coding_reviews_fable_template.csv",
            output_name=FABLE_WORKBOOK,
            blind=False,
        ),
        build_submission_workbook(
            package_dir,
            review_template="profile_coding_reviews_blind_template.csv",
            output_name=BLIND_WORKBOOK,
            blind=True,
        ),
    ]
    return outputs


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--package-dir",
        type=Path,
        default=DEFAULT_PACKAGE_DIR,
        help="Directory containing the machine-readable worksheet templates.",
    )
    args = parser.parse_args()
    outputs = build_all(args.package_dir.resolve())
    for output in outputs:
        print(f"Built {output}")


if __name__ == "__main__":
    main()
