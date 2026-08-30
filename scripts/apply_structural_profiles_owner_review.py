#!/usr/bin/env python3
"""Apply the adopted owner-review workbook without altering reconciliation inputs."""

from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PACKAGE = ROOT / "research" / "structural-profiles-pilot" / "reconciliation"
WORKBOOK_PATH = PACKAGE / "owner_exception_review_v1.xlsx"

ALLOWED_DISPOSITIONS = {
    "prefer_seed",
    "prefer_independent",
    "preserve_disagreement",
    "needs_domain_review",
    "needs_better_evidence",
}
UNRESOLVED_DISPOSITIONS = {
    "preserve_disagreement",
    "needs_domain_review",
    "needs_better_evidence",
}
EXPECTED_DISPOSITION_COUNTS = Counter(
    {
        "prefer_seed": 9,
        "prefer_independent": 7,
        "needs_better_evidence": 4,
        "needs_domain_review": 2,
        "preserve_disagreement": 1,
    }
)
S5_CONVENTION_CHOICE = "include_reasonably_foreseeable_escaped_consequences"
POST_RECONCILIATION_ROUTE = "clarify_S5_then_targeted_S5_adjudication"
PRIORITY_STAGE_IDS = {
    "experiment_selection",
    "plasma_control",
    "materials_qualification",
    "tritium_and_fuel_cycle",
    "blankets",
    "commissioning",
    "reliability_demonstration",
    "licensing",
    "grid_integration",
}
NO_COMPARABLE_CASE_BASIS = {
    "sp-0024": (
        "The submitted rationale states that pilot-scale tritium/fuel-cycle "
        "performance has not been demonstrated in a directly comparable case."
    ),
    "sp-0025": (
        "The submitted rationale states that integrated blanket performance "
        "cannot yet be observed in a fusion-relevant test environment."
    ),
    "sp-0031": (
        "The adopted owner rationale states that no directly comparable "
        "fusion-grid case exists."
    ),
}


class OwnerReviewBuildError(ValueError):
    """Raised when the adopted workbook or a protected input violates the contract."""


def require(condition: bool, message: str) -> None:
    if not condition:
        raise OwnerReviewBuildError(message)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def extract_owner_decisions() -> list[dict[str, str]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=False, data_only=False)
    require(
        workbook.sheetnames
        == [
            "OWNER_SUMMARY",
            "OWNER_EXCEPTIONS",
            "CROSS_CUTTING_S5",
            "PM_RECOMMENDATIONS",
        ],
        "Adopted workbook sheet set/order drifted.",
    )

    summary_note = workbook["OWNER_SUMMARY"]["A18"].value
    require(
        isinstance(summary_note, str) and summary_note.startswith("PM recommendation added:"),
        "Adopted workbook is missing the clearly labeled PM summary note.",
    )

    cross = workbook["CROSS_CUTTING_S5"]
    require(
        cross["B10"].value == S5_CONVENTION_CHOICE,
        "Adopted workbook has the wrong S5 convention choice.",
    )
    require(
        cross["G10"].value == POST_RECONCILIATION_ROUTE,
        "Adopted workbook has the wrong post-reconciliation route.",
    )

    owner_rows = read_csv(PACKAGE / "owner_exceptions_v1.csv")
    require(len(owner_rows) == 23, "Owner exception routing must remain 23 rows.")
    owner_sheet = workbook["OWNER_EXCEPTIONS"]
    workbook_rows = [
        row
        for row in owner_sheet.iter_rows(min_row=4, max_row=26, values_only=True)
        if row[0]
    ]
    require(len(workbook_rows) == 23, "Adopted workbook must contain 23 decisions.")

    decisions: list[dict[str, str]] = []
    for source, workbook_row in zip(owner_rows, workbook_rows):
        exception_id = str(workbook_row[0])
        disposition = str(workbook_row[21] or "")
        rationale = str(workbook_row[22] or "")
        require(
            exception_id == source["exception_id"],
            f"Workbook/exception order drift at {exception_id}.",
        )
        require(
            disposition in ALLOWED_DISPOSITIONS,
            f"{exception_id} has no single allowed owner disposition.",
        )
        require(bool(rationale.strip()), f"{exception_id} has no owner rationale.")
        decisions.append(
            {
                "exception_id": exception_id,
                "profile_id": source["profile_id"],
                "stage_id": source["stage_id"],
                "s_dimension": source["s_dimension"],
                "owner_disposition": disposition,
                "owner_rationale": rationale,
                "decision_state": (
                    "unresolved"
                    if disposition in UNRESOLVED_DISPOSITIONS
                    else "owner_preference_recorded"
                ),
                "selected_s_value": "",
            }
        )

    require(
        Counter(row["owner_disposition"] for row in decisions)
        == EXPECTED_DISPOSITION_COUNTS,
        "Owner disposition counts do not match the adopted workbook summary.",
    )
    return decisions


def build_targeted_s5_backlog(
    decisions: list[dict[str, str]],
) -> list[dict[str, Any]]:
    audit_rows = read_csv(PACKAGE / "comparison_audit_v1.csv")
    s5_rows = [
        row
        for row in audit_rows
        if row["s_dimension"] == "S5" and int(row["numeric_difference"] or 0) > 0
    ]
    require(len(s5_rows) == 19, "Targeted S5 backlog must contain 19 affected rows.")
    decisions_by_key = {
        (row["profile_id"], row["s_dimension"]): row for row in decisions
    }

    backlog: list[dict[str, Any]] = []
    for row in s5_rows:
        decision = decisions_by_key.get((row["profile_id"], "S5"))
        backlog.append(
            {
                "profile_id": row["profile_id"],
                "stage_id": row["stage_id"],
                "workflow": row["workflow"],
                "sector": row["sector"],
                "pathway_id": row["pathway_id"],
                "seed_review_id": row["seed_review_id"],
                "seed_s5": row["seed_value"],
                "independent_review_id": row["independent_review_id"],
                "independent_s5": row["independent_value"],
                "numeric_difference": row["numeric_difference"],
                "comparison_status": row["comparison_status"],
                "owner_exception_id": decision["exception_id"] if decision else "",
                "owner_disposition": decision["owner_disposition"] if decision else "",
                "owner_rationale": decision["owner_rationale"] if decision else "",
                "adjudication_status": (
                    "owner_disposition_recorded_pending_targeted_s5_adjudication"
                    if decision
                    else "audit_trail_retained_pending_targeted_s5_adjudication"
                ),
                "selected_s5": "",
                "post_reconciliation_route": POST_RECONCILIATION_ROUTE,
            }
        )

    require(
        sum(bool(row["owner_exception_id"]) for row in backlog) == 7,
        "Targeted S5 backlog must carry seven owner-routed decisions.",
    )
    return backlog


def build_evidence_backlog(
    fusion_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    backlog: list[dict[str, str]] = [
        {
            "backlog_id": "eg-fusion-pack-inventory",
            "profile_id": "",
            "stage_id": "all_fusion_profiles",
            "workflow": "All 18 fusion profiles",
            "priority": "package",
            "evidence_gap_type": "internal_fusion_test_pack_not_yet_banked",
            "gap_basis": (
                "The PM workbook identifies The Fusion Test evidence pack as an "
                "internal input, but its claims have not been mapped to profile rows "
                "or banked in the canonical source register."
            ),
            "evidence_locator": (
                "The Fusion Test — Evidence Pack for Frontier Is Not Fate "
                "(research cut-off 2026-08-12)"
            ),
            "source_id": "",
            "next_action": (
                "Inventory and verify the pack, map claims to profiles, and assign "
                "source IDs only in the later source-banking gate."
            ),
            "status": "open",
        }
    ]

    for row in fusion_rows:
        priority = "priority" if row["stage_id"] in PRIORITY_STAGE_IDS else "routine"
        common = {
            "profile_id": row["profile_id"],
            "stage_id": row["stage_id"],
            "workflow": row["workflow"],
            "priority": priority,
            "evidence_locator": "",
            "source_id": "",
            "status": "open",
        }
        backlog.append(
            {
                "backlog_id": f"eg-{row['profile_id']}-source-id",
                **common,
                "evidence_gap_type": "canonical_source_id_missing",
                "gap_basis": (
                    "Both pinned model submissions leave source_ids blank for this "
                    "profile; agreement or an owner preference is not evidence."
                ),
                "next_action": (
                    "Bank verified supporting sources later; do not invent or promote "
                    "a source ID in this correction pass."
                ),
            }
        )
        backlog.append(
            {
                "backlog_id": f"eg-{row['profile_id']}-expert",
                **common,
                "evidence_gap_type": "regulatory_or_domain_expert_required",
                "gap_basis": (
                    "Every frozen fusion profile is routed for routine "
                    "domain-informed review; licensing additionally requires a "
                    "fusion-regulatory reviewer."
                ),
                "next_action": (
                    "Assign a named fusion-domain reviewer; assign a regulatory "
                    "reviewer for licensing."
                ),
            }
        )
        if row["profile_id"] in NO_COMPARABLE_CASE_BASIS:
            backlog.append(
                {
                    "backlog_id": f"eg-{row['profile_id']}-observed-case",
                    **common,
                    "evidence_gap_type": "no_directly_comparable_observed_case_exists",
                    "gap_basis": NO_COMPARABLE_CASE_BASIS[row["profile_id"]],
                    "next_action": (
                        "Preserve the limitation and use domain-reviewed analogies or "
                        "explicitly bounded inference; do not fabricate an observed case."
                    ),
                }
            )
    return backlog


def build_fusion_brief(fusion_rows: list[dict[str, str]]) -> str:
    priority = [row for row in fusion_rows if row["stage_id"] in PRIORITY_STAGE_IDS]
    routine = [row for row in fusion_rows if row["stage_id"] not in PRIORITY_STAGE_IDS]
    require(len(priority) == 9 and len(routine) == 9, "Fusion priority split drifted.")

    lines = [
        "# Fusion domain-review brief — v1",
        "",
        "## Route and boundary",
        "",
        "All 18 frozen magnetic-confinement fusion profiles remain queued for routine domain-informed review. This review is a post-reconciliation evidence and scope gate: it does not approve or canonicalize a profile, populate WP2, promote a source, or replace either model submission.",
        "",
        "The review must preserve the original seed and independent S-values and rationales. Owner dispositions are review instructions, not canonical profile rows. S5-only work follows `clarify_S5_then_targeted_S5_adjudication`; S1-S4 are not reopened in that pass.",
        "",
        "## Priority review profiles",
        "",
        "These nine profiles receive the first domain-review slots because the owner review identifies load-bearing evidence, scope, assurance, or no-comparable-case questions.",
        "",
        "| Profile | Stage | Lifecycle | Current route |",
        "| --- | --- | --- | --- |",
    ]
    for row in priority:
        lines.append(
            f"| `{row['profile_id']}` | {row['workflow']} | `{row['lifecycle_phase']}` | priority domain review |"
        )
    lines.extend(
        [
            "",
            "Priority topics are experiment selection, plasma control, materials qualification, tritium/fuel cycle, blankets, commissioning, reliability demonstration, licensing, and grid integration.",
            "",
            "## Remaining routine review profiles",
            "",
            "| Profile | Stage | Lifecycle | Current route |",
            "| --- | --- | --- | --- |",
        ]
    )
    for row in routine:
        lines.append(
            f"| `{row['profile_id']}` | {row['workflow']} | `{row['lifecycle_phase']}` | routine domain review |"
        )
    lines.extend(
        [
            "",
            "## Evidence-gap routing",
            "",
            "The companion `evidence_gap_backlog_v1.csv` keeps four states separate:",
            "",
            "- evidence identified in the internal Fusion Test pack but not yet inventoried, verified, mapped, or banked;",
            "- a canonical source ID is missing from the model submissions;",
            "- a regulatory or fusion-domain expert is required; and",
            "- no directly comparable observed case exists for the stated profile claim.",
            "",
            "A single profile can carry more than one state. Blank source IDs remain missing. This PR does not convert PM evidence leads into source-register entries and does not infer a source ID from a URL or document title.",
            "",
            "## Next gate",
            "",
            "The next gate is source inventory and banking followed by named domain review. Targeted S5 adjudication can then use the clarified convention and the preserved 19-row audit trail. Until those steps are complete, no fusion profile is approved or canonical.",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> None:
    decisions = extract_owner_decisions()
    decisions_fields = list(decisions[0])
    write_csv(PACKAGE / "owner_decisions_v1.csv", decisions_fields, decisions)

    s5_backlog = build_targeted_s5_backlog(decisions)
    write_csv(
        PACKAGE / "targeted_s5_adjudication_backlog_v1.csv",
        list(s5_backlog[0]),
        s5_backlog,
    )

    fusion_rows = read_csv(PACKAGE / "fusion_domain_review_queue_v1.csv")
    require(len(fusion_rows) == 18, "Fusion domain-review queue must remain 18 rows.")
    evidence_backlog = build_evidence_backlog(fusion_rows)
    write_csv(
        PACKAGE / "evidence_gap_backlog_v1.csv",
        list(evidence_backlog[0]),
        evidence_backlog,
    )
    (PACKAGE / "fusion_domain_review_brief_v1.md").write_text(
        build_fusion_brief(fusion_rows), encoding="utf-8"
    )
    print(
        "Applied owner review: 23 decisions, 19 targeted S5 rows, "
        f"{len(evidence_backlog)} evidence backlog rows, 18 fusion reviews."
    )


if __name__ == "__main__":
    main()
