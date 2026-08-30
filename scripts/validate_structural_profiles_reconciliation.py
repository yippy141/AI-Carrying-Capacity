#!/usr/bin/env python3
"""Validate the issue #31 Structural Profiles reconciliation package."""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PACKAGE = ROOT / "research" / "structural-profiles-pilot" / "reconciliation"
WORKSHEET = ROOT / "research" / "structural-profiles-pilot" / "worksheet"

EXPECTED_HASHES = {
    "seed_submission_v1.csv": "9723a3c3a006b701fed3d000f77d89379610bc51f85eff3555b696078708e675",
    "independent_submission_v1.csv": "4176afa878a850d7155a95772884bec72756353440b588f30e5414a37acdb973",
}
EXPECTED_HEADS = {
    "seed_submission_id": "pr-33@3bc7e1d3e16d507fe7374ae66a3af6eace519ca4",
    "independent_submission_id": "pr-32@86dbff4aab694d413a33d3c4a8b0d28047d73c2f",
}
DIMENSIONS = ["S1", "S2", "S3", "S4", "S5"]
STATUSES = [
    "exact_agreement",
    "one_point_difference",
    "difference_ge_2",
    "missing_seed",
    "missing_independent",
    "missing_both",
]
EXPECTED_COUNTS = {
    "exact_agreement": 96,
    "one_point_difference": 57,
    "difference_ge_2": 2,
    "missing_seed": 0,
    "missing_independent": 0,
    "missing_both": 0,
}
EXPECTED_FILES = {
    "RECONCILIATION_NOTE.md",
    "seed_submission_v1.csv",
    "independent_submission_v1.csv",
    "comparison_audit_v1.csv",
    "reconciliation_report_v1.json",
    "owner_exceptions_v1.csv",
    "owner_exception_review_v1.xlsx",
    "fusion_domain_review_queue_v1.csv",
}
EXPECTED_SUBMISSION_KEYS = {
    "review_id",
    "coder_role",
    "coder_name",
    "coder_model",
    "reasoning_effort",
    "value",
    "rationale",
    "source_ids",
    "confidence",
}


class ReconciliationValidationError(ValueError):
    """Raised when the reconciliation package violates its contract."""


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ReconciliationValidationError(message)


def _schema_type_matches(instance: object, expected: str) -> bool:
    return {
        "object": isinstance(instance, dict),
        "array": isinstance(instance, list),
        "string": isinstance(instance, str),
        "integer": isinstance(instance, int) and not isinstance(instance, bool),
        "boolean": isinstance(instance, bool),
        "null": instance is None,
    }.get(expected, False)


def validate_json_schema(
    instance: object, schema: dict, root_schema: dict, path: str = "$"
) -> None:
    if "$ref" in schema:
        reference = schema["$ref"]
        require(reference.startswith("#/$defs/"), f"Unsupported schema reference {reference}.")
        definition = root_schema["$defs"][reference.rsplit("/", 1)[1]]
        validate_json_schema(instance, definition, root_schema, path)
        return
    if "anyOf" in schema:
        matched = False
        for candidate in schema["anyOf"]:
            try:
                validate_json_schema(instance, candidate, root_schema, path)
                matched = True
                break
            except ReconciliationValidationError:
                continue
        require(matched, f"{path} does not match any allowed schema branch.")
        return
    if "const" in schema:
        require(instance == schema["const"], f"{path} violates const.")
    if "enum" in schema:
        require(instance in schema["enum"], f"{path} is not an allowed enum value.")
    if "type" in schema:
        allowed_types = schema["type"] if isinstance(schema["type"], list) else [schema["type"]]
        require(
            any(_schema_type_matches(instance, expected) for expected in allowed_types),
            f"{path} has the wrong JSON type; expected {allowed_types}.",
        )
    if isinstance(instance, dict):
        required = schema.get("required", [])
        missing = [key for key in required if key not in instance]
        require(not missing, f"{path} is missing required keys {missing}.")
        properties = schema.get("properties", {})
        if schema.get("additionalProperties") is False:
            extra = sorted(set(instance) - set(properties))
            require(not extra, f"{path} has additional properties {extra}.")
        for key, value in instance.items():
            if key in properties:
                validate_json_schema(value, properties[key], root_schema, f"{path}.{key}")
    if isinstance(instance, list):
        if "minItems" in schema:
            require(len(instance) >= schema["minItems"], f"{path} has too few items.")
        if "maxItems" in schema:
            require(len(instance) <= schema["maxItems"], f"{path} has too many items.")
        if schema.get("uniqueItems"):
            canonical = [json.dumps(item, sort_keys=True, ensure_ascii=False) for item in instance]
            require(len(canonical) == len(set(canonical)), f"{path} items are not unique.")
        if "items" in schema:
            for index, value in enumerate(instance):
                validate_json_schema(value, schema["items"], root_schema, f"{path}[{index}]")
    if isinstance(instance, str):
        if "minLength" in schema:
            require(len(instance) >= schema["minLength"], f"{path} is too short.")
        if "pattern" in schema:
            require(re.search(schema["pattern"], instance) is not None, f"{path} fails pattern.")
        if schema.get("format") == "date-time":
            try:
                datetime.fromisoformat(instance.replace("Z", "+00:00"))
            except ValueError as exc:
                raise ReconciliationValidationError(
                    f"{path} is not a valid date-time."
                ) from exc
    if isinstance(instance, int) and not isinstance(instance, bool):
        if "minimum" in schema:
            require(instance >= schema["minimum"], f"{path} is below minimum.")
        if "maximum" in schema:
            require(instance <= schema["maximum"], f"{path} is above maximum.")


def validate_schema_normalization() -> None:
    schema_text = (WORKSHEET / "exception_report.schema.json").read_text(
        encoding="utf-8"
    )
    schema = json.loads(schema_text)
    require("fable" not in schema_text.lower(), "Exception schema retains Fable labels.")
    metadata_required = schema["properties"]["metadata"]["required"]
    require(
        "seed_submission_id" in metadata_required,
        "Exception schema must require seed_submission_id.",
    )
    statuses = schema["$defs"]["comparisonStatus"]["enum"]
    require(statuses == STATUSES, "Exception schema comparison statuses drifted.")
    dispositions = schema["$defs"]["exception"]["properties"][
        "owner_disposition"
    ]["enum"]
    require(
        dispositions
        == [
            "prefer_seed",
            "prefer_independent",
            "preserve_disagreement",
            "needs_domain_review",
            "needs_better_evidence",
            None,
        ],
        "Exception schema owner dispositions drifted.",
    )


def validate_raw_submissions() -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    rows_by_label = {}
    for file_name, expected_hash in EXPECTED_HASHES.items():
        data = (PACKAGE / file_name).read_bytes()
        require(
            hashlib.sha256(data).hexdigest() == expected_hash,
            f"{file_name} is not the immutable pinned copy.",
        )
        rows = read_csv(PACKAGE / file_name)
        require(len(rows) == 31, f"{file_name} must contain 31 profiles.")
        require(
            len({row["profile_id"] for row in rows}) == 31,
            f"{file_name} profile IDs must be unique.",
        )
        require(
            all(not row["source_ids"] for row in rows),
            f"{file_name} systematic blank-source count drifted.",
        )
        rows_by_label[file_name] = rows
    seed = rows_by_label["seed_submission_v1.csv"]
    independent = rows_by_label["independent_submission_v1.csv"]
    require(
        [row["profile_id"] for row in seed]
        == [row["profile_id"] for row in independent],
        "Pinned submission profile order differs.",
    )
    return seed, independent


def validate_report(
    seed_rows: list[dict[str, str]], independent_rows: list[dict[str, str]]
) -> tuple[dict, list[dict[str, str]]]:
    report = json.loads(
        (PACKAGE / "reconciliation_report_v1.json").read_text(encoding="utf-8")
    )
    schema = json.loads(
        (WORKSHEET / "exception_report.schema.json").read_text(encoding="utf-8")
    )
    validate_json_schema(report, schema, schema)
    require(
        set(report)
        == {
            "metadata",
            "comparison_audit",
            "summary_counts",
            "exceptions",
            "routine_domain_reviews",
        },
        "Reconciliation report top-level keys drifted.",
    )
    for key, value in EXPECTED_HEADS.items():
        require(report["metadata"][key] == value, f"Report metadata {key} drifted.")
    require(
        report["metadata"]["comparison_rule"]
        == "compare_seed_and_independent_by_profile_id_and_s_dimension",
        "Report uses the wrong comparison rule.",
    )
    require(
        report["summary_counts"]
        == {**EXPECTED_COUNTS, "total_comparisons": 155},
        "Report comparison counts drifted.",
    )

    audit = report["comparison_audit"]
    require(len(audit) == 155, "Report must contain 155 audit rows.")
    expected_keys = [
        (row["profile_id"], dimension)
        for row in seed_rows
        for dimension in DIMENSIONS
    ]
    require(
        [(row["profile_id"], row["s_dimension"]) for row in audit] == expected_keys,
        "Audit profile/dimension order drifted.",
    )
    require(
        Counter(row["comparison_status"] for row in audit) == Counter(EXPECTED_COUNTS),
        "Audit status counts drifted.",
    )

    seed_by_id = {row["profile_id"]: row for row in seed_rows}
    independent_by_id = {row["profile_id"]: row for row in independent_rows}
    for row in audit:
        profile_id = row["profile_id"]
        dimension = row["s_dimension"]
        require(set(row["seed"]) == EXPECTED_SUBMISSION_KEYS, "Seed view keys drifted.")
        require(
            set(row["independent"]) == EXPECTED_SUBMISSION_KEYS,
            "Independent view keys drifted.",
        )
        require(
            row["seed"]["value"] == int(seed_by_id[profile_id][dimension]),
            f"Seed value drift at {profile_id} {dimension}.",
        )
        require(
            row["independent"]["value"]
            == int(independent_by_id[profile_id][dimension]),
            f"Independent value drift at {profile_id} {dimension}.",
        )
        require(
            row["seed"]["rationale"] == seed_by_id[profile_id]["rationale"],
            f"Seed rationale drift at {profile_id} {dimension}.",
        )
        require(
            row["independent"]["rationale"]
            == independent_by_id[profile_id]["rationale"],
            f"Independent rationale drift at {profile_id} {dimension}.",
        )
        require(
            (row["seed"]["coder_name"], row["seed"]["coder_model"])
            == ("Claude Code", "claude-opus-5"),
            f"Seed identity drift at {profile_id} {dimension}.",
        )
        require(
            (
                row["independent"]["coder_name"],
                row["independent"]["coder_model"],
                row["independent"]["reasoning_effort"],
            )
            == ("Codex", "gpt-5.6", "extra_high"),
            f"Independent identity drift at {profile_id} {dimension}.",
        )
        require(
            row["seed"]["source_ids"] == []
            and row["independent"]["source_ids"] == [],
            f"Systematic source gap drift at {profile_id} {dimension}.",
        )

    exceptions = report["exceptions"]
    require(len(exceptions) == 23, "Owner exception count must remain 23.")
    require(
        all(
            row["owner_review_required"]
            and row["owner_disposition"] is None
            and row["owner_rationale"] is None
            and row["status"] == "pending_owner"
            for row in exceptions
        ),
        "Owner decisions must remain blank and pending.",
    )
    audit_exception_keys = {
        (row["profile_id"], row["s_dimension"])
        for row in audit
        if row["owner_review_required"]
    }
    require(
        audit_exception_keys
        == {(row["profile_id"], row["s_dimension"]) for row in exceptions},
        "Owner exception subset does not match the audit routing flags.",
    )

    fusion = report["routine_domain_reviews"]
    require(len(fusion) == 18, "Routine fusion queue must contain 18 profiles.")
    require(
        [row["profile_id"] for row in fusion]
        == [f"sp-{index:04d}" for index in range(14, 32)],
        "Routine fusion queue profile set/order drifted.",
    )
    require(
        all(
            row["owner_decision_required"] is False
            and row["domain_reviewer"] is None
            and row["notes"] is None
            and row["seed_coder_name"] == "Claude Code"
            and row["seed_coder_model"] == "claude-opus-5"
            and row["independent_coder_name"] == "Codex"
            and row["independent_coder_model"] == "gpt-5.6"
            and row["independent_reasoning_effort"] == "extra_high"
            for row in fusion
        ),
        "Routine domain-review provenance or blank-review fields drifted.",
    )
    return report, read_csv(PACKAGE / "comparison_audit_v1.csv")


def validate_flat_outputs(report: dict, audit_csv: list[dict[str, str]]) -> None:
    require(len(audit_csv) == 155, "comparison_audit_v1.csv must contain 155 rows.")
    require(
        Counter(row["comparison_status"] for row in audit_csv)
        == Counter(EXPECTED_COUNTS),
        "Flattened audit counts drifted.",
    )
    owner_rows = read_csv(PACKAGE / "owner_exceptions_v1.csv")
    require(len(owner_rows) == 23, "owner_exceptions_v1.csv must contain 23 rows.")
    require(
        all(
            not row["owner_disposition"]
            and not row["owner_rationale"]
            and row["status"] == "pending_owner"
            for row in owner_rows
        ),
        "Owner CSV contains a prefilled decision.",
    )
    require(
        {(row["profile_id"], row["s_dimension"]) for row in owner_rows}
        == {
            (row["profile_id"], row["s_dimension"])
            for row in report["exceptions"]
        },
        "Owner CSV does not match report exceptions.",
    )
    fusion_rows = read_csv(PACKAGE / "fusion_domain_review_queue_v1.csv")
    require(len(fusion_rows) == 18, "Fusion CSV must contain exactly 18 profiles.")
    require(
        all(
            row["owner_decision_required"] == "false"
            and not row["domain_reviewer"]
            and not row["notes"]
            for row in fusion_rows
        ),
        "Fusion queue must remain a blank routine domain-review queue.",
    )


def validate_owner_workbook() -> None:
    workbook = load_workbook(
        PACKAGE / "owner_exception_review_v1.xlsx", read_only=False, data_only=False
    )
    require(
        workbook.sheetnames
        == ["OWNER_SUMMARY", "OWNER_EXCEPTIONS", "CROSS_CUTTING_S5"],
        "Owner workbook sheet set/order drifted.",
    )
    owner_sheet = workbook["OWNER_EXCEPTIONS"]
    headers = [cell.value for cell in owner_sheet[3]]
    require(
        headers[-2:] == ["owner disposition", "owner rationale"],
        "Owner workbook editable columns drifted.",
    )
    owner_data = [row for row in owner_sheet.iter_rows(min_row=4, values_only=True) if row[0]]
    require(len(owner_data) == 23, "Owner workbook must show exactly 23 exceptions.")
    require(
        all(row[-2] is None and row[-1] is None for row in owner_data),
        "Owner workbook contains a prefilled row decision.",
    )
    cross = workbook["CROSS_CUTTING_S5"]
    require(
        cross["B9"].value is None and cross["G9"].value is None,
        "S5 convention choice or correction route must remain blank.",
    )
    affected = [row for row in cross.iter_rows(min_row=14, values_only=True) if row[0]]
    require(len(affected) == 19, "Owner workbook must identify all 19 S5 differences.")
    formulas = [
        cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    ]
    require(formulas == ["=SUM(B4:B9)"], "Owner workbook formula set drifted.")


def validate_note() -> None:
    note = (PACKAGE / "RECONCILIATION_NOTE.md").read_text(encoding="utf-8")
    required_phrases = [
        "Owner-routed exceptions: **23**",
        "exact_agreement` | 96",
        "one_point_difference` | 57",
        "difference_ge_2` | 2",
        "blank canonical `source_ids` on 31 of 31 profile rows",
        "locally contained errors only",
        "reasonably foreseeable consequences",
        "exactly 18 profiles",
        "does not select, approve, canonicalize, or implement any profile row",
    ]
    for phrase in required_phrases:
        require(phrase in note, f"Reconciliation note is missing: {phrase}")


def validate() -> None:
    actual_files = {path.name for path in PACKAGE.iterdir() if path.is_file()}
    require(
        actual_files == EXPECTED_FILES,
        f"Unexpected reconciliation package files: {sorted(actual_files ^ EXPECTED_FILES)}",
    )
    validate_schema_normalization()
    seed_rows, independent_rows = validate_raw_submissions()
    report, audit_csv = validate_report(seed_rows, independent_rows)
    validate_flat_outputs(report, audit_csv)
    validate_owner_workbook()
    validate_note()
    print(
        "Structural Profiles reconciliation validation passed: "
        "155 comparisons, 23 owner exceptions, 18 fusion reviews."
    )


if __name__ == "__main__":
    validate()
